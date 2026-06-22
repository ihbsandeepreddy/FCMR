"""Deterministic NBFC audit sampling engine — 10M-row scale.

Implements all four criteria from the audit sampling specification:
  1. Full population coverage — stratified by Product × Status × Branch.
  2. Targeted exception coverage — up to MAX_PER_EXCEPTION rows per class,
     newest first; every detected class guaranteed in the sample.
  3. Unbiased random sampling from the clean population.
  4. Recency bias as a weight [0.25, 1.0] — never a hard filter.

All heavy paths (exception detection, recency weight, strata) are fully
vectorised Polars expressions; the only Python-level loop is the per-class
sample build (≤ len(EXCEPTION_RULES) iterations, each O(1) after filtering).

SEED=42, stable sort, no time-dependent randomness → reproducible.
"""

from __future__ import annotations

import random
from datetime import date, datetime, timedelta
from pathlib import Path

import polars as pl

# ── constants ────────────────────────────────────────────────────────────────
SEED: int = 42
MAX_PER_EXCEPTION: int = 10
TARGET_PCT: float = 0.005      # 0.5 % of clean population
TARGET_MIN: int = 50           # minimum random fill rows
IQR_MULTIPLIER: float = 1.5
TODAY: date = date.today()
EARLIEST_VALID_DATE: date = date(1990, 1, 1)

# Every exception class — order is audit-significance priority (used only for
# display; multi-flag means every matching rule fires independently).
EXCEPTION_RULES: list[str] = [
    "DUPLICATE_LAN",
    "DUPLICATE_CUSTOMER",
    "NULL_MANDATORY_FIELD",
    "INVALID_SANCTIONED_AMT",
    "OUTSTANDING_EXCEEDS_SANCTIONED",
    "INVALID_INTEREST_RATE",
    "INVALID_TENURE",
    "FUTURE_LOAN_DATE",
    "IMPLAUSIBLY_OLD_DATE",
    "CLOSED_WITH_OUTSTANDING",
    "STATUS_DPD_MISMATCH",
    "KYC_ISSUE",
    "OUTLIER_SANCTIONED_AMT",
    "OUTLIER_OUTSTANDING_AMT",
    "OUTLIER_INTEREST_RATE",
]


# ── synthetic data generator ─────────────────────────────────────────────────

def generate_synthetic_loan_master(n_rows: int = 10_000_000) -> pl.DataFrame:
    """Generate a realistic synthetic NBFC loan customer master.

    Base data is clean; exceptions are injected in controlled proportions so
    every exception class is represented and the total exception rate is ~12 %.

    Realistic distributions:
      • Loan status: 60 % Active, 30 % Closed, 7 % NPA, 3 % Written-off
      • KYC: 85 % Complete, 5 % Expired, 5 % Pending, 5 % Missing
      • DPD: 0 for Closed; weighted toward 0 for Active; ≥ 90 for NPA/WO
      • Outstanding ≤ Sanctioned in base data (mismatch only via injection)
    """
    rng = random.Random(SEED)

    n_products = 50
    n_branches = 200
    # 70 % unique-customer coverage → ~1.4 loans per customer on average
    n_unique_customers = int(n_rows * 0.70)

    base_date = datetime(2019, 1, 1)
    date_pool = [
        (base_date + timedelta(days=rng.randint(0, 365 * 5))).strftime("%Y-%m-%d")
        for _ in range(n_rows)
    ]
    status_pool = rng.choices(
        ["Active", "Closed", "NPA", "Written-off"],
        weights=[60, 30, 7, 3],
        k=n_rows,
    )
    sanctioned_pool = [rng.randint(50_000, 10_000_000) for _ in range(n_rows)]

    # Outstanding ≤ Sanctioned in clean base; Closed → 0; NPA/WO → large
    def _outstanding(status: str, sanctioned: int) -> int:
        if status == "Closed":
            return 0
        if status in ("NPA", "Written-off"):
            return rng.randint(int(sanctioned * 0.4), sanctioned)
        return rng.randint(0, sanctioned)

    outstanding_pool = [_outstanding(s, a) for s, a in zip(status_pool, sanctioned_pool)]

    def _dpd(status: str) -> int:
        if status == "Closed":
            return 0
        if status == "NPA":
            return rng.choice([90, 120, 150, 180, 270])
        if status == "Written-off":
            return rng.choice([360, 540, 720])
        return rng.choices([0, 30, 60], weights=[70, 20, 10])[0]

    records = {
        "LAN":            [f"LAN{i:010d}" for i in range(n_rows)],
        "CUSTOMER_ID":    [f"CUST{rng.randint(1000, n_unique_customers):08d}" for _ in range(n_rows)],
        "LOAN_DATE":      date_pool,
        "SANCTIONED_AMT": sanctioned_pool,
        "OUTSTANDING_AMT":outstanding_pool,
        "INTEREST_RATE":  [round(rng.uniform(7.5, 22.0), 2) for _ in range(n_rows)],
        "TENURE_MONTHS":  [rng.choice([12, 24, 36, 48, 60, 84]) for _ in range(n_rows)],
        "LOAN_STATUS":    status_pool,
        "DPD":            [_dpd(s) for s in status_pool],
        "PRODUCT":        [f"PROD{rng.randint(1, n_products):03d}" for _ in range(n_rows)],
        "BRANCH":         [f"BR{rng.randint(1, n_branches):04d}" for _ in range(n_rows)],
        "KYC_STATUS":     rng.choices(
            ["Complete", "Expired", "Pending", "Missing"],
            weights=[85, 5, 5, 5],
            k=n_rows,
        ),
    }

    df = pl.DataFrame(records)

    # ── inject targeted exceptions so every class appears ──────────────────
    # Roughly 1–2 % per class → ~12 % total exception rate.
    rows = df.to_dicts()
    n_inject = max(50, int(n_rows * 0.01))   # ~1 % per injection batch
    available = list(range(n_rows))
    rng.shuffle(available)
    chunks = [available[i * n_inject:(i + 1) * n_inject] for i in range(15)]

    injection_ops = [
        # (chunk_idx, field, value_fn)
        (0,  "LAN",            lambda i, r: f"LAN{(i // 2) * 2:010d}"),  # paired dups
        (1,  "SANCTIONED_AMT", lambda i, r: 0),
        (2,  "OUTSTANDING_AMT",lambda i, r: r["SANCTIONED_AMT"] * 2),
        (3,  "INTEREST_RATE",  lambda i, r: 0.0),
        (4,  "TENURE_MONTHS",  lambda i, r: 0),
        (5,  "LOAN_DATE",      lambda i, r: "2027-01-01"),           # future
        (6,  "LOAN_DATE",      lambda i, r: "1985-06-15"),           # pre-1990
        (7,  "LOAN_STATUS",    lambda i, r: "Closed"),               # closed + POS > 0 …
        (7,  "OUTSTANDING_AMT",lambda i, r: max(r["SANCTIONED_AMT"] // 2, 1)),
        (8,  "LOAN_STATUS",    lambda i, r: "NPA"),                  # NPA + DPD = 0
        (8,  "DPD",            lambda i, r: 0),
        (9,  "KYC_STATUS",     lambda i, r: rng.choice(["Missing", "Expired", "Pending"])),
        (10, "INTEREST_RATE",  lambda i, r: round(rng.uniform(50, 120), 2)),  # outlier
        (11, "SANCTIONED_AMT", lambda i, r: rng.randint(900_000_000, 999_999_999)),  # outlier
        (12, "OUTSTANDING_AMT",lambda i, r: rng.randint(800_000_000, 999_999_999)),  # outlier
        (13, "CUSTOMER_ID",    lambda i, r: f"CUST{1000 + (i % (n_inject // 2)):08d}"),  # dup customers
        (14, "LAN",            lambda i, r: None),                   # null mandatory
    ]

    for chunk_idx, field, val_fn in injection_ops:
        if chunk_idx >= len(chunks):
            continue
        for idx in chunks[chunk_idx]:
            rows[idx][field] = val_fn(idx, rows[idx])

    return pl.DataFrame(rows).with_row_index("ROW_ID")  # ROW_ID added fresh here


# ── exception detection (fully vectorised) ───────────────────────────────────

def _iqr_bounds(series: pl.Series) -> tuple[float, float]:
    """Return (lower, upper) IQR-based outlier bounds for a numeric series."""
    q1 = series.quantile(0.25) or 0.0
    q3 = series.quantile(0.75) or 0.0
    iqr = q3 - q1
    return q1 - IQR_MULTIPLIER * iqr, q3 + IQR_MULTIPLIER * iqr


def detect_exceptions(df: pl.DataFrame) -> pl.DataFrame:
    """Add one boolean column per exception class and a multi-flag string.

    Returns df with:
      _exc_<RULE>  — bool, True if this row violates this rule
      EXCEPTION_FLAGS — semicolon-joined list of all active rules ('' if clean)
      IS_EXCEPTION    — 1 if any rule fires, else 0

    All detection is vectorised (no Python row loops).
    """
    # ── duplicate sets (cross-row; must do before adding columns) ──────────
    dup_lans = (
        df["LAN"]
        .drop_nulls()
        .value_counts()
        .filter(pl.col("count") > 1)["LAN"]
    )
    dup_customers = (
        df["CUSTOMER_ID"]
        .drop_nulls()
        .value_counts()
        .filter(pl.col("count") > 1)["CUSTOMER_ID"]
    )

    # ── IQR bounds ──────────────────────────────────────────────────────────
    lo_san, hi_san = _iqr_bounds(df["SANCTIONED_AMT"].drop_nulls())
    lo_out, hi_out = _iqr_bounds(df["OUTSTANDING_AMT"].drop_nulls())
    lo_rate, hi_rate = _iqr_bounds(df["INTEREST_RATE"].drop_nulls())

    # ── parse date column once ──────────────────────────────────────────────
    parsed_date = pl.col("LOAN_DATE").str.to_date(format="%Y-%m-%d", strict=False)

    # ── build all boolean exception columns in one with_columns call ────────
    df = df.with_columns(
        # G2 — duplicate LAN
        (pl.col("LAN").is_null() | pl.col("LAN").is_in(dup_lans))
        .alias("_exc_DUPLICATE_LAN"),

        # G3 — null mandatory fields (LAN, CUSTOMER_ID, LOAN_DATE, SANCTIONED_AMT, LOAN_STATUS)
        (
            pl.col("LAN").is_null()
            | pl.col("CUSTOMER_ID").is_null()
            | pl.col("LOAN_DATE").is_null()
            | pl.col("SANCTIONED_AMT").is_null()
            | pl.col("LOAN_STATUS").is_null()
        ).alias("_exc_NULL_MANDATORY_FIELD"),

        # Invalid sanctioned amount
        (pl.col("SANCTIONED_AMT").is_null() | (pl.col("SANCTIONED_AMT") <= 0))
        .alias("_exc_INVALID_SANCTIONED_AMT"),

        # Outstanding > Sanctioned
        (
            pl.col("OUTSTANDING_AMT").is_not_null()
            & pl.col("SANCTIONED_AMT").is_not_null()
            & (pl.col("OUTSTANDING_AMT") > pl.col("SANCTIONED_AMT"))
        ).alias("_exc_OUTSTANDING_EXCEEDS_SANCTIONED"),

        # Invalid interest rate
        (
            pl.col("INTEREST_RATE").is_null()
            | (pl.col("INTEREST_RATE") <= 0)
            | (pl.col("INTEREST_RATE") > 36)
        ).alias("_exc_INVALID_INTEREST_RATE"),

        # Invalid tenure
        (pl.col("TENURE_MONTHS").is_null() | (pl.col("TENURE_MONTHS") <= 0))
        .alias("_exc_INVALID_TENURE"),

        # G4 — future loan date
        (parsed_date.is_not_null() & (parsed_date > pl.lit(TODAY)))
        .alias("_exc_FUTURE_LOAN_DATE"),

        # G4 — impossibly old date (< 1990)
        (parsed_date.is_not_null() & (parsed_date < pl.lit(EARLIEST_VALID_DATE)))
        .alias("_exc_IMPLAUSIBLY_OLD_DATE"),

        # Closed loan with positive outstanding
        (
            (pl.col("LOAN_STATUS") == "Closed")
            & pl.col("OUTSTANDING_AMT").is_not_null()
            & (pl.col("OUTSTANDING_AMT") > 0)
        ).alias("_exc_CLOSED_WITH_OUTSTANDING"),

        # NPA/Written-off with DPD = 0
        (
            pl.col("LOAN_STATUS").is_in(["NPA", "Written-off"])
            & pl.col("DPD").is_not_null()
            & (pl.col("DPD") == 0)
        ).alias("_exc_STATUS_DPD_MISMATCH"),

        # KYC not complete
        pl.col("KYC_STATUS").is_in(["Missing", "Expired", "Pending"])
        .alias("_exc_KYC_ISSUE"),

        # G5 — IQR outliers
        (
            pl.col("SANCTIONED_AMT").is_not_null()
            & ((pl.col("SANCTIONED_AMT") < lo_san) | (pl.col("SANCTIONED_AMT") > hi_san))
        ).alias("_exc_OUTLIER_SANCTIONED_AMT"),

        (
            pl.col("OUTSTANDING_AMT").is_not_null()
            & ((pl.col("OUTSTANDING_AMT") < lo_out) | (pl.col("OUTSTANDING_AMT") > hi_out))
        ).alias("_exc_OUTLIER_OUTSTANDING_AMT"),

        (
            pl.col("INTEREST_RATE").is_not_null()
            & ((pl.col("INTEREST_RATE") < lo_rate) | (pl.col("INTEREST_RATE") > hi_rate))
        ).alias("_exc_OUTLIER_INTEREST_RATE"),
    )

    # Duplicate-customer flag requires a separate with_columns call after
    # the LAN duplicate (both need the base df["CUSTOMER_ID"] column intact).
    df = df.with_columns(
        pl.col("CUSTOMER_ID").is_in(dup_customers).alias("_exc_DUPLICATE_CUSTOMER")
    )

    # ── build multi-flag string per row (G1 — multi-flag) ──────────────────
    exc_cols = [f"_exc_{r}" for r in EXCEPTION_RULES]
    # For each rule: emit the rule name if True, null otherwise; concat_str skips nulls.
    flag_parts = [
        pl.when(pl.col(c)).then(pl.lit(c[5:])).otherwise(pl.lit(None))
        for c in exc_cols
    ]
    df = df.with_columns(
        pl.concat_str(flag_parts, separator=";", ignore_nulls=True)
        .alias("EXCEPTION_FLAGS"),
    )

    # IS_EXCEPTION — any rule fires
    df = df.with_columns(
        pl.any_horizontal([pl.col(c) for c in exc_cols])
        .cast(pl.Int8)
        .alias("IS_EXCEPTION")
    )

    return df


# ── recency weight (vectorised) ───────────────────────────────────────────────

def compute_recency_weight(df: pl.DataFrame) -> pl.DataFrame:
    """Assign recency_weight in [0.25, 1.0] using LOAN_DATE.

    Fallback: if fewer than 2 distinct dates exist, use numeric LAN suffix
    as the sequence proxy (per spec: "fall back to sequential LAN order only
    if LAN is demonstrably sequential").
    """
    parsed = df["LOAN_DATE"].str.to_date(format="%Y-%m-%d", strict=False)
    n_distinct = parsed.drop_nulls().n_unique()

    if n_distinct >= 2:
        # Vectorised date-based rank
        ts = parsed.cast(pl.Int64, strict=False)   # epoch days; null → null
        ts_filled = ts.fill_null(ts.median())
        min_ts, max_ts = ts_filled.min(), ts_filled.max()
        spread = (max_ts - min_ts) or 1
        weights = 0.25 + 0.75 * ((ts_filled - min_ts) / spread)
        recency_basis = "LOAN_DATE"
    else:
        # Fallback: use trailing numeric part of LAN as sequence
        lan_num = (
            df["LAN"]
            .str.extract(r"(\d+)$", 1)
            .cast(pl.Float64, strict=False)
            .fill_null(0.0)
        )
        mn, mx = lan_num.min(), lan_num.max()
        spread = (mx - mn) or 1
        weights = 0.25 + 0.75 * ((lan_num - mn) / spread)
        recency_basis = "LAN_SEQUENCE (fallback — fewer than 2 distinct dates)"

    return df.with_columns(
        weights.cast(pl.Float64).alias("RECENCY_WEIGHT")
    ), recency_basis


# ── sampling phases ───────────────────────────────────────────────────────────

def _sample_exceptions(df: pl.DataFrame) -> tuple[pl.DataFrame, dict[str, int]]:
    """Phase 1 — per-class exception sample.

    For each exception class present in df:
      - Sort by RECENCY_WEIGHT DESC (newest first)
      - Take up to MAX_PER_EXCEPTION rows
      - De-duplicate by ROW_ID: a multi-flagged row counts once in the output
        but is credited to each class whose quota it fills.

    Returns (sample_df, {class: n_selected}).
    """
    exc_cols = [f"_exc_{r}" for r in EXCEPTION_RULES]
    selected_ids: set[int] = set()
    class_rows: list[pl.DataFrame] = []
    class_counts: dict[str, int] = {}

    for rule, col in zip(EXCEPTION_RULES, exc_cols):
        if col not in df.columns:
            class_counts[rule] = 0
            continue

        candidates = (
            df.filter(pl.col(col))
            .sort("RECENCY_WEIGHT", descending=True)
        )
        if candidates.is_empty():
            class_counts[rule] = 0
            continue

        # Fill up to MAX_PER_EXCEPTION, preferring unselected rows first
        new_rows = candidates.filter(~pl.col("ROW_ID").is_in(list(selected_ids))).head(MAX_PER_EXCEPTION)
        already_counted = max(0, MAX_PER_EXCEPTION - len(new_rows))
        # Credit already-selected multi-flag rows toward this class quota
        already_in_class = candidates.filter(pl.col("ROW_ID").is_in(list(selected_ids))).head(already_counted)

        n_new = len(new_rows)
        n_credited = len(already_in_class)
        class_counts[rule] = n_new + n_credited

        if n_new > 0:
            labeled = new_rows.with_columns(
                pl.lit(f"EXCEPTION:{rule}").alias("SELECTION_REASON")
            )
            class_rows.append(labeled)
            selected_ids.update(new_rows["ROW_ID"].to_list())

    if class_rows:
        sample = pl.concat(class_rows, how="diagonal_relaxed")
    else:
        sample = pl.DataFrame()

    return sample, class_counts


def _sample_strata(df_clean: pl.DataFrame) -> pl.DataFrame:
    """Phase 2 — guarantee ≥ 1 row from every non-empty stratum.

    Strata key: PRODUCT × LOAN_STATUS × BRANCH (collapsed to available dims).
    Within each stratum, picks the row with highest RECENCY_WEIGHT.
    Fully vectorised: one group_by call.
    """
    dims = [c for c in ["PRODUCT", "LOAN_STATUS", "BRANCH"] if c in df_clean.columns]
    if not dims:
        return pl.DataFrame()

    strata = (
        df_clean
        .sort("RECENCY_WEIGHT", descending=True)
        .group_by(dims)
        .head(1)
    )
    return strata.with_columns(
        (
            pl.lit("STRATA_COVERAGE:")
            + pl.concat_str([pl.col(d) for d in dims], separator="|")
        ).alias("SELECTION_REASON")
    )


def _sample_random(
    df_clean: pl.DataFrame,
    exclude_ids: set[int],
    n_target: int,
) -> pl.DataFrame:
    """Phase 3 — recency-weighted random fill from remaining clean rows.

    Uses numpy if available for O(n) weighted choice without replacement;
    falls back to random.choices (with-replacement, then dedup) if numpy absent.
    """
    if n_target <= 0:
        return pl.DataFrame()

    pool = df_clean.filter(~pl.col("ROW_ID").is_in(list(exclude_ids)))
    if pool.is_empty():
        return pl.DataFrame()

    k = min(n_target, len(pool))
    weights = pool["RECENCY_WEIGHT"].to_list()
    total_w = sum(weights)
    probs = [w / total_w for w in weights]

    try:
        import numpy as np
        rng = np.random.default_rng(SEED)
        indices = rng.choice(len(pool), size=k, replace=False, p=probs).tolist()
    except ImportError:
        random.seed(SEED)
        indices = list({i for i in random.choices(range(len(pool)), weights=probs, k=k * 3)})[:k]

    return pool[indices].with_columns(pl.lit("RANDOM").alias("SELECTION_REASON"))


# ── main entry point ──────────────────────────────────────────────────────────

def run_audit_sampling(df: pl.DataFrame) -> tuple[pl.DataFrame, dict]:
    """Run the full audit sampling pipeline.

    Returns:
        selected_sample — DataFrame with appended columns:
            SELECTION_REASON, EXCEPTION_FLAGS, RECENCY_WEIGHT
        summary — dict with counts, per-class stats, and assumptions
    """
    n_pop = len(df)

    # 1. Detect exceptions (vectorised, multi-flag)
    df = detect_exceptions(df)

    # 2. Recency weight (vectorised)
    df, recency_basis = compute_recency_weight(df)

    # 3. Split clean / exception
    df_exc = df.filter(pl.col("IS_EXCEPTION") == 1)
    df_clean = df.filter(pl.col("IS_EXCEPTION") == 0)
    n_exc = len(df_exc)
    n_clean = len(df_clean)

    # Per-class totals (for summary)
    exc_totals: dict[str, int] = {}
    for rule in EXCEPTION_RULES:
        col = f"_exc_{rule}"
        exc_totals[rule] = int(df.filter(pl.col(col)).height) if col in df.columns else 0

    # Phase 1 — exception sample
    sample_exc, class_counts = _sample_exceptions(df_exc)

    # Phase 2 — strata coverage
    sample_strata = _sample_strata(df_clean)
    n_strata = len(sample_strata)

    # Phase 3 — random fill
    used_ids: set[int] = set()
    if sample_exc.height > 0 and "ROW_ID" in sample_exc.columns:
        used_ids.update(sample_exc["ROW_ID"].to_list())
    if sample_strata.height > 0 and "ROW_ID" in sample_strata.columns:
        used_ids.update(sample_strata["ROW_ID"].to_list())

    n_random_target = max(int(n_clean * TARGET_PCT), TARGET_MIN) - n_strata
    sample_random = _sample_random(df_clean, used_ids, n_random_target)

    # Combine — SELECTION_REASON column present in all three phases
    parts = [s for s in [sample_exc, sample_strata, sample_random] if s.height > 0]
    if parts:
        selected = pl.concat(parts, how="diagonal_relaxed")
    else:
        selected = pl.DataFrame()

    # Drop internal boolean columns from output (keep EXCEPTION_FLAGS instead)
    exc_bool_cols = [f"_exc_{r}" for r in EXCEPTION_RULES if f"_exc_{r}" in selected.columns]
    if exc_bool_cols:
        selected = selected.drop(exc_bool_cols)

    # ── summary ──────────────────────────────────────────────────────────────
    n_classes_found = sum(1 for v in exc_totals.values() if v > 0)
    n_classes_sampled = sum(1 for v in class_counts.values() if v > 0)

    summary: dict = {
        "population_total": n_pop,
        "clean_records": n_clean,
        "exception_records": n_exc,
        "exception_classes_found": n_classes_found,
        "exception_classes_sampled": n_classes_sampled,
        "exceptions_sampled": len(sample_exc),
        "strata_covered": n_strata,
        "random_sampled": len(sample_random),
        "total_sample_size": len(selected),
        "sample_pct": round(len(selected) / n_pop * 100, 3) if n_pop > 0 else 0.0,
        "recency_basis": recency_basis,
        "per_class": {
            rule: {"found": exc_totals.get(rule, 0), "sampled": class_counts.get(rule, 0)}
            for rule in EXCEPTION_RULES
        },
        "assumptions": _build_assumptions(df),
    }

    return selected, summary


def _build_assumptions(df: pl.DataFrame) -> list[str]:
    """Document any assumptions/fallbacks for the audit trail."""
    notes: list[str] = []
    missing = [c for c in ["LOAN_DATE", "SANCTIONED_AMT", "OUTSTANDING_AMT", "INTEREST_RATE", "TENURE_MONTHS", "DPD", "KYC_STATUS"] if c not in df.columns]
    if missing:
        notes.append(f"Missing columns (rules skipped): {', '.join(missing)}")
    null_dates = df["LOAN_DATE"].is_null().sum() if "LOAN_DATE" in df.columns else 0
    if null_dates:
        notes.append(f"{null_dates:,} rows had null LOAN_DATE; recency weight used median date as fill.")
    notes.append(f"IQR multiplier: {IQR_MULTIPLIER}; outlier bounds computed on non-null values.")
    notes.append(f"Sampling seed: {SEED} (reproducible).")
    notes.append(f"Max per exception class: {MAX_PER_EXCEPTION}.")
    notes.append(f"Random sample target: max({TARGET_PCT*100:.1f}% of clean, {TARGET_MIN} rows).")
    return notes


# ── Excel export ───────────────────────────────────────────────────────────────

def export_to_excel(sample: pl.DataFrame, summary: dict, output_path: Path) -> None:
    """Export the audit sample to a multi-sheet Excel workbook.

    Sheets:
      1. Audit Sample    — all selected rows with SELECTION_REASON / EXCEPTION_FLAGS
      2. Exception Sample — rows selected for exception coverage only
      3. Strata Sample   — rows selected for strata coverage only
      4. Random Sample   — rows from the random phase
      5. Summary         — population / sample counts, per-class stats, assumptions
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="C65D2E")   # terracotta
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ALT_FILL    = PatternFill("solid", fgColor="FDF6F0")

    def _write_sheet(ws, df_sheet: pl.DataFrame, title: str) -> None:
        ws.title = title
        if df_sheet.is_empty():
            ws.append(["(no data)"])
            return
        headers = df_sheet.columns
        ws.append(list(headers))
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        for r_idx, row in enumerate(df_sheet.to_dicts(), start=2):
            ws.append([row.get(h) for h in headers])
            if r_idx % 2 == 0:
                for cell in ws[r_idx]:
                    cell.fill = ALT_FILL
        for col_idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        ws.freeze_panes = "A2"

    wb = Workbook()

    # Sheet 1 — full sample
    _write_sheet(wb.active, sample, "Audit Sample")

    # Sheets 2-4 — by phase
    for title, reason_prefix in [
        ("Exception Sample", "EXCEPTION:"),
        ("Strata Sample",    "STRATA_COVERAGE:"),
        ("Random Sample",    "RANDOM"),
    ]:
        if "SELECTION_REASON" in sample.columns:
            sub = sample.filter(pl.col("SELECTION_REASON").str.starts_with(reason_prefix))
        else:
            sub = pl.DataFrame()
        _write_sheet(wb.create_sheet(), sub, title)

    # Sheet 5 — summary
    ws_sum = wb.create_sheet("Summary")
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 18

    top_stats = [
        ("Population total",           summary["population_total"]),
        ("Clean records",              summary["clean_records"]),
        ("Exception records",          summary["exception_records"]),
        ("Exception classes found",    summary["exception_classes_found"]),
        ("Exception classes sampled",  summary["exception_classes_sampled"]),
        ("Exceptions sampled",         summary["exceptions_sampled"]),
        ("Strata covered",             summary["strata_covered"]),
        ("Random sampled",             summary["random_sampled"]),
        ("Total sample size",          summary["total_sample_size"]),
        ("Sample % of population",     f"{summary['sample_pct']:.3f}%"),
        ("Recency basis",              summary["recency_basis"]),
    ]
    ws_sum.append(["Metric", "Value"])
    for cell in ws_sum[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in top_stats:
        ws_sum.append(list(row))

    ws_sum.append([])
    ws_sum.append(["Exception class", "Found", "Sampled"])
    header_row = ws_sum.max_row
    for cell in ws_sum[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for rule, counts in summary["per_class"].items():
        ws_sum.append([rule, counts["found"], counts["sampled"]])

    ws_sum.append([])
    ws_sum.append(["Assumption / note", ""])
    for note in summary["assumptions"]:
        ws_sum.append([note, ""])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"[OK] Exported to {output_path}")
