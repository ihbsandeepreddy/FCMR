"""EAD Portfolio Analytics — Ind AS 109 / ECL report suite.

Operates on a merged Polars DataFrame of all EAD uploads for an engagement.
All compute functions are pure Polars (vectorised, no Python loops).
Missing columns are skipped gracefully so the suite works even when a field
was not mapped or was not present in the source files.

5 primary analytics filter to the **last month** using the ``business_date``
column (max value = most recent snapshot):
  - UCID → LAN Count        (ucid_lan)
  - State-wise EAD           (state_ead)
  - Disbursements in Period  (disbursement)  — uses engagement period dates
  - Product-wise EAD         (product_ead)
  - Product × Stage × DPD   (product_stage_dpd)
"""

from __future__ import annotations

import datetime as _dt
from pathlib import Path

import polars as pl

from fcmr_core.analytics.ead_summary import (
    generate_collateral_coverage,
    generate_data_quality_summary_ead,
    generate_dpd_risk_distribution,
    generate_portfolio_concentration,
    generate_provision_coverage,
    generate_sanction_disbursement_variance,
    generate_stage_distribution,
    generate_writeoff_recovery,
)
from fcmr_core.logging_setup import get_logger

logger = get_logger("processing")

# ---------------------------------------------------------------------------
# Column maps
# ---------------------------------------------------------------------------

# canonical field name → display name used in pivot output columns
PIVOT_DIMENSION_CANONICAL: dict[str, str] = {
    "entity": "Entity",
    "source_system_id": "SourcesystemID",
    "scheme_id": "SchemeID",
    "scheme_name": "SchemeName",
    "branch_id": "BranchID",
    "branch_name": "BranchName",
    "state": "State",
    "sbu": "SBU",
    "sub_sbu": "SubSBU",
    "channel_code": "ChannelCode",
    "channel_desc": "ChannelCodeDesc",
    "loan_status": "LoanStatus",
    "stage": "Stage",
    "dpd_bucket": "DPDBucketing",
    "written_off": "WrittenOff",
}

# raw column names not in ead_files.yaml canonical schema — pass through as-is
PIVOT_DIMENSION_RAW: list[str] = [
    "System",
    "ECL",
    "WriteOffSecuSelldown",
    "ProductID",
    "ProductName",
    "FVTPL",
    "FuturePoS_GL_ID",
    "FuturePOS_GL_Des",
    "Principle_DRS_GL_ID",
    "Principle_DRS_GL_Des",
    "INT_DRS_GL_ID",
    "INT_DRS_GL_DESC",
    "Accrued_Interest_GL_ID",
    "Accrued_Interest_GL_DESC",
    "Redemption_Premium_Debtors_GL_ID",
    "Redemption_Premium_Debtors_GL_DESC",
    "Accrued_Redemption_Premium_GL_ID",
    "Accrued_Redemption_Premium_GL_DESC",
    "Prepaid_GL_ID",
    "Prepaid_GL_DESC",
]

# canonical field name → pivot output display name (SUM columns)
PIVOT_MEASURE_CANONICAL: dict[str, str] = {
    "future_pos": "Sum(FuturePOS)",
    "outstanding_principal": "Sum(DrsPOS)",
    "outstanding_interest": "Sum(DrsInt)",
    "accrued_interest": "Sum(AccruedInterest)",
    "overdue_charges": "Sum(OverDueCharges)",
    "gross_book_value": "Sum(Gross_Loans_and_Advances)",
    "ead": "Sum(EAD)",
    "mob": "Sum(MOB)",
    "collateral_value": "Sum(National_Asset_Value)",
    "covered_portion": "Sum(Covered_Portion)",
    "uncovered_portion": "Sum(Uncovered_Portion)",
    "existing_provision": "Sum(Provsion_as_per_Policy)",
    "additional_provision": "Sum(AdditionalProvision)",
    "total_provision": "Sum(Total_Provision)",
    "sanction_amount": "Sum(SANCTIONAMOUNT)",
    "disbursed_amount": "Sum(DISBURSEDAMOUNT)",
    "loan_id": "Count(AGREEMENTID_R)",  # COUNT, not SUM
    "principal_paid": "Sum(PrincipalPaid)",
    "interest_paid": "Sum(InterestPaid)",
    "emi_amount": "Sum(EMI Amount)",
}

# raw column names for measures not in the canonical schema
PIVOT_MEASURE_RAW: list[str] = [
    "Prepaid",
    "Debtors_Redemption_Premium",
    "AccruedRedemptionPremium",
    "TotalDebtors",
    "zero_90_days_interest",
    "zero_90_days_interest_Hist",
    "zero_90_int_Final",
    "Other Charges Paid",
    "PrinAdvance",
    "CBC Unpaid",
    "ODC Unpaid",
    "CBC",
    "ODC",
    "CBC Paid",
    "ODC Paid",
    "PROCESSINGFEES",
    "Unbanked",
]

# DPD bucket strings that imply > 90 days overdue
_HIGH_DPD_PATTERNS = (
    "91",
    "121",
    "151",
    "181",
    "211",
    "241",
    "271",
    "361",
    "npa",
    "sub",
    "doubtful",
    "loss",
)

# Loan status values indicating account closure / write-off
_CLOSED_STATUSES = {"closed", "written off", "writeoff", "wo", "write off", "settled", "foreclosed"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _safe_float(df: pl.DataFrame, col: str) -> pl.Series:
    """Cast a column to Float64, returning nulls for non-numeric values."""
    return df[col].cast(pl.Float64, strict=False)


def _is_high_dpd(series: pl.Series) -> pl.Series:
    """Boolean mask: DPD bucket implies > 90 days overdue."""
    lower = series.cast(pl.Utf8, strict=False).str.to_lowercase().fill_null("")
    return lower.map_elements(
        lambda v: any(p in v for p in _HIGH_DPD_PATTERNS),
        return_dtype=pl.Boolean,
    )


def _available_dims(df: pl.DataFrame) -> list[str]:
    """Return canonical dimension column names present in df, in display-name order."""
    cols = []
    for canonical in PIVOT_DIMENSION_CANONICAL:
        if canonical in df.columns:
            cols.append(canonical)
    for raw in PIVOT_DIMENSION_RAW:
        if raw in df.columns:
            cols.append(raw)
    return cols


def _filter_last_month(df: pl.DataFrame) -> pl.DataFrame:
    """Filter to rows where business_date == max(business_date).

    Uses the canonical ``business_date`` column added to the EAD schema.
    If the column is absent or unparseable, returns df unchanged.
    """
    if "business_date" not in df.columns:
        return df
    # Try to parse as Date for correct chronological sorting
    parsed = df["business_date"].cast(pl.Date, strict=False)
    non_null = parsed.drop_nulls()
    if non_null.len() > 0:
        max_val = non_null.max()
        return df.filter(parsed == max_val)
    # Fallback: string comparison (works if dates are already in ISO YYYY-MM-DD)
    max_str = df["business_date"].drop_nulls().max()
    if max_str is None:
        return df
    return df.filter(pl.col("business_date") == max_str)


# ---------------------------------------------------------------------------
# PRIMARY ANALYTICS 1 — UCID → LAN Count (last month)
# ---------------------------------------------------------------------------


def generate_ucid_lan_count(df: pl.DataFrame) -> pl.DataFrame:
    """Count distinct LANs per UCID for the last business_date month."""
    df = _filter_last_month(df)
    if "ucid" not in df.columns:
        return pl.DataFrame(
            {"note": ["ucid column not available — map the UCID column in EAD schema"]}
        )
    if "loan_id" not in df.columns:
        return pl.DataFrame({"note": ["loan_id column not available"]})

    agg: list[pl.Expr] = [pl.col("loan_id").count().alias("LAN Count")]
    if "ead" in df.columns:
        agg.append(pl.col("ead").cast(pl.Float64, strict=False).sum().alias("Total EAD"))
    if "total_provision" in df.columns:
        agg.append(
            pl.col("total_provision").cast(pl.Float64, strict=False).sum().alias("Total Provision")
        )

    result = df.group_by("ucid").agg(agg).rename({"ucid": "UCID"})
    return result.sort("LAN Count", descending=True)


# ---------------------------------------------------------------------------
# PRIMARY ANALYTICS 2 — State-wise EAD Balance (last month)
# ---------------------------------------------------------------------------


def generate_state_ead(df: pl.DataFrame) -> pl.DataFrame:
    """State-wise EAD, loan count and provision for the last business_date month."""
    df = _filter_last_month(df)
    if "state" not in df.columns:
        return pl.DataFrame({"note": ["state column not available"]})

    agg: list[pl.Expr] = []
    if "loan_id" in df.columns:
        agg.append(pl.col("loan_id").count().alias("Loan Count"))
    if "ead" in df.columns:
        agg.append(pl.col("ead").cast(pl.Float64, strict=False).sum().alias("Total EAD"))
    if "total_provision" in df.columns:
        agg.append(
            pl.col("total_provision").cast(pl.Float64, strict=False).sum().alias("Total Provision")
        )
    if not agg:
        agg = [pl.len().alias("Loan Count")]

    result = df.group_by("state").agg(agg).rename({"state": "State"})

    if "Total EAD" in result.columns and "Total Provision" in result.columns:
        result = result.with_columns(
            (pl.col("Total Provision") / pl.col("Total EAD") * 100).round(2).alias("Coverage %")
        )

    sort_col = "Total EAD" if "Total EAD" in result.columns else "State"
    return result.sort(sort_col, descending=("EAD" in sort_col))


# ---------------------------------------------------------------------------
# PRIMARY ANALYTICS 3 — Disbursements in Audit Period
# ---------------------------------------------------------------------------


def generate_disbursement_summary(
    df: pl.DataFrame,
    period_from: str | None = None,
    period_to: str | None = None,
) -> pl.DataFrame:
    """Disbursements where disbursement_date falls within the engagement period."""
    if "disbursement_date" not in df.columns:
        return pl.DataFrame({"note": ["disbursement_date column not available"]})

    df = df.with_columns(
        pl.col("disbursement_date").cast(pl.Date, strict=False).alias("_disb_date")
    )

    if period_from:
        try:
            from_dt = _dt.date.fromisoformat(period_from)
            df = df.filter(pl.col("_disb_date") >= pl.lit(from_dt))
        except (ValueError, TypeError):
            pass

    if period_to:
        try:
            to_dt = _dt.date.fromisoformat(period_to)
            df = df.filter(pl.col("_disb_date") <= pl.lit(to_dt))
        except (ValueError, TypeError):
            pass

    df = df.filter(pl.col("_disb_date").is_not_null())

    if df.is_empty():
        period_note = (
            f" for period {period_from or '?'} to {period_to or '?'}"
            if (period_from or period_to)
            else ""
        )
        return pl.DataFrame({"note": [f"No disbursements found{period_note}"]})

    df = df.with_columns(pl.col("_disb_date").dt.strftime("%Y-%m").alias("Disbursement Month"))

    group_cols = ["Disbursement Month"] + [
        c for c in ["scheme_id", "scheme_name", "state"] if c in df.columns
    ]

    agg: list[pl.Expr] = []
    if "loan_id" in df.columns:
        agg.append(pl.col("loan_id").count().alias("New Loans"))
    if "disbursed_amount" in df.columns:
        agg.append(
            pl.col("disbursed_amount").cast(pl.Float64, strict=False).sum().alias("Total Disbursed")
        )
    if "sanction_amount" in df.columns:
        agg.append(
            pl.col("sanction_amount").cast(pl.Float64, strict=False).sum().alias("Total Sanctioned")
        )
    if not agg:
        agg = [pl.len().alias("New Loans")]

    rename_map = {"scheme_id": "Product ID", "scheme_name": "Product Name", "state": "State"}
    result = (
        df.group_by(group_cols)
        .agg(agg)
        .rename({k: v for k, v in rename_map.items() if k in group_cols})
    )

    sort_cols = ["Disbursement Month"] + [
        rename_map.get(c, c) for c in ["scheme_id", "scheme_name", "state"] if c in group_cols
    ]
    valid_sort = [c for c in sort_cols if c in result.columns]
    return result.sort(valid_sort) if valid_sort else result


# ---------------------------------------------------------------------------
# PRIMARY ANALYTICS 4 — Product-wise EAD Balance (last month)
# ---------------------------------------------------------------------------


def generate_product_ead(df: pl.DataFrame) -> pl.DataFrame:
    """Product-wise (scheme_id / ProductID) EAD for the last business_date month."""
    df = _filter_last_month(df)
    group_cols = [c for c in ["scheme_id", "scheme_name"] if c in df.columns]
    if not group_cols:
        return pl.DataFrame(
            {"note": ["scheme_id / scheme_name (ProductID / ProductName) columns not available"]}
        )

    agg: list[pl.Expr] = []
    if "loan_id" in df.columns:
        agg.append(pl.col("loan_id").count().alias("Loan Count"))
    if "ead" in df.columns:
        agg.append(pl.col("ead").cast(pl.Float64, strict=False).sum().alias("Total EAD"))
    if "total_provision" in df.columns:
        agg.append(
            pl.col("total_provision").cast(pl.Float64, strict=False).sum().alias("Total Provision")
        )
    if "outstanding_principal" in df.columns:
        agg.append(
            pl.col("outstanding_principal")
            .cast(pl.Float64, strict=False)
            .sum()
            .alias("Outstanding Principal")
        )
    if not agg:
        agg = [pl.len().alias("Loan Count")]

    rename_map = {"scheme_id": "Product ID", "scheme_name": "Product Name"}
    result = (
        df.group_by(group_cols)
        .agg(agg)
        .rename({k: v for k, v in rename_map.items() if k in group_cols})
    )

    if "Total EAD" in result.columns and "Total Provision" in result.columns:
        result = result.with_columns(
            (pl.col("Total Provision") / pl.col("Total EAD") * 100).round(2).alias("Coverage %")
        )

    sort_col = "Total EAD" if "Total EAD" in result.columns else result.columns[0]
    return result.sort(sort_col, descending=("EAD" in sort_col))


# ---------------------------------------------------------------------------
# PRIMARY ANALYTICS 5 — Product × Stage × DPD (last month)
# ---------------------------------------------------------------------------


def generate_product_stage_dpd(df: pl.DataFrame) -> pl.DataFrame:
    """Stage × DPD breakdown by product for the last business_date month."""
    df = _filter_last_month(df)
    group_cols = [c for c in ["scheme_id", "scheme_name", "stage", "dpd_bucket"] if c in df.columns]
    if len(group_cols) < 2:
        return pl.DataFrame(
            {"note": ["Need at least scheme_id/scheme_name + stage/dpd_bucket columns"]}
        )

    agg: list[pl.Expr] = []
    if "loan_id" in df.columns:
        agg.append(pl.col("loan_id").count().alias("Loan Count"))
    if "ead" in df.columns:
        agg.append(pl.col("ead").cast(pl.Float64, strict=False).sum().alias("Total EAD"))
    if "total_provision" in df.columns:
        agg.append(
            pl.col("total_provision").cast(pl.Float64, strict=False).sum().alias("Total Provision")
        )
    if not agg:
        agg = [pl.len().alias("Loan Count")]

    rename_map = {
        "scheme_id": "Product ID",
        "scheme_name": "Product Name",
        "stage": "Stage",
        "dpd_bucket": "DPD Bucket",
    }
    result = (
        df.group_by(group_cols)
        .agg(agg)
        .rename({k: v for k, v in rename_map.items() if k in group_cols})
    )

    # Flag: DPD > 90 but Stage = 1
    if "DPD Bucket" in result.columns and "Stage" in result.columns:
        high_dpd = _is_high_dpd(result["DPD Bucket"])
        stage_1 = result["Stage"].cast(pl.Utf8, strict=False).str.strip_chars() == "1"
        result = result.with_columns((high_dpd & stage_1).alias("DPD-Stage Mismatch"))

    sort_cols = [
        c for c in ["Product Name", "Product ID", "Stage", "DPD Bucket"] if c in result.columns
    ]
    return result.sort(sort_cols) if sort_cols else result


# ---------------------------------------------------------------------------
# 6. Master pivot report (all months)
# ---------------------------------------------------------------------------


def generate_pivot_report(df: pl.DataFrame) -> pl.DataFrame:
    """GROUP BY all available dimension columns, SUM/COUNT all measure columns."""
    dim_cols = _available_dims(df)

    agg_exprs: list[pl.Expr] = []

    for canonical, display in PIVOT_MEASURE_CANONICAL.items():
        if canonical not in df.columns:
            continue
        col_f = _safe_float(df, canonical)
        tmp_col = f"__tmp_{canonical}"
        df = df.with_columns(col_f.alias(tmp_col))
        if canonical == "loan_id":
            agg_exprs.append(pl.col(tmp_col).count().alias(display))
        else:
            agg_exprs.append(pl.col(tmp_col).sum().alias(display))

    for raw in PIVOT_MEASURE_RAW:
        if raw not in df.columns:
            continue
        col_f = _safe_float(df, raw)
        tmp_col = f"__tmp_raw_{raw.replace(' ', '_')}"
        df = df.with_columns(col_f.alias(tmp_col))
        agg_exprs.append(pl.col(tmp_col).sum().alias(f"Sum({raw})"))

    if not agg_exprs:
        return pl.DataFrame()

    if dim_cols:
        result = df.group_by(dim_cols).agg(agg_exprs).sort(dim_cols)
    else:
        result = df.select(agg_exprs)

    result = result.select([c for c in result.columns if not c.startswith("__tmp_")])
    return result


# ---------------------------------------------------------------------------
# 7. Stage-wise ECL summary (all months)
# ---------------------------------------------------------------------------


def generate_stage_summary(df: pl.DataFrame) -> pl.DataFrame:
    if "stage" not in df.columns:
        return pl.DataFrame({"note": ["stage column not available"]})

    agg: list[pl.Expr] = []
    if "ead" in df.columns:
        agg.append(pl.col("ead").cast(pl.Float64, strict=False).sum().alias("Total EAD"))
    if "total_provision" in df.columns:
        agg.append(
            pl.col("total_provision").cast(pl.Float64, strict=False).sum().alias("Total Provision")
        )
    if "existing_provision" in df.columns:
        agg.append(
            pl.col("existing_provision")
            .cast(pl.Float64, strict=False)
            .sum()
            .alias("Provision as per Policy")
        )
    if "additional_provision" in df.columns:
        agg.append(
            pl.col("additional_provision")
            .cast(pl.Float64, strict=False)
            .sum()
            .alias("Additional Provision")
        )
    if "loan_id" in df.columns:
        agg.append(pl.col("loan_id").count().alias("Loan Count"))

    if not agg:
        return pl.DataFrame({"stage": df["stage"].unique()})

    result = df.group_by("stage").agg(agg).sort("stage")

    if "Total EAD" in result.columns and "Total Provision" in result.columns:
        result = result.with_columns(
            (pl.col("Total Provision") / pl.col("Total EAD") * 100).round(2).alias("Coverage %")
        )

    return result


# ---------------------------------------------------------------------------
# 8. DPD bucket summary (all months)
# ---------------------------------------------------------------------------


def generate_dpd_summary(df: pl.DataFrame) -> pl.DataFrame:
    if "dpd_bucket" not in df.columns:
        return pl.DataFrame({"note": ["dpd_bucket column not available"]})

    agg: list[pl.Expr] = []
    if "loan_id" in df.columns:
        agg.append(pl.col("loan_id").count().alias("Loan Count"))
    if "ead" in df.columns:
        agg.append(pl.col("ead").cast(pl.Float64, strict=False).sum().alias("Total EAD"))
    if "outstanding_principal" in df.columns:
        agg.append(
            pl.col("outstanding_principal")
            .cast(pl.Float64, strict=False)
            .sum()
            .alias("Outstanding Principal")
        )
    if "total_provision" in df.columns:
        agg.append(
            pl.col("total_provision").cast(pl.Float64, strict=False).sum().alias("Total Provision")
        )

    if not agg:
        return pl.DataFrame({"DPD Bucket": df["dpd_bucket"].unique()})

    return (
        df.group_by("dpd_bucket").agg(agg).rename({"dpd_bucket": "DPD Bucket"}).sort("DPD Bucket")
    )


# ---------------------------------------------------------------------------
# 9. Secured vs unsecured (all months)
# ---------------------------------------------------------------------------


def generate_security_summary(df: pl.DataFrame) -> pl.DataFrame:
    if "stage" not in df.columns:
        return pl.DataFrame({"note": ["stage column not available"]})

    agg: list[pl.Expr] = []
    if "covered_portion" in df.columns:
        agg.append(
            pl.col("covered_portion")
            .cast(pl.Float64, strict=False)
            .sum()
            .alias("Covered (Secured)")
        )
    if "uncovered_portion" in df.columns:
        agg.append(
            pl.col("uncovered_portion")
            .cast(pl.Float64, strict=False)
            .sum()
            .alias("Uncovered (Unsecured)")
        )
    if "ead" in df.columns:
        agg.append(pl.col("ead").cast(pl.Float64, strict=False).sum().alias("Total EAD"))
    if not agg:
        return pl.DataFrame({"note": ["covered_portion / uncovered_portion columns not available"]})

    result = df.group_by("stage").agg(agg).rename({"stage": "Stage"}).sort("Stage")

    if "Covered (Secured)" in result.columns and "Total EAD" in result.columns:
        result = result.with_columns(
            (pl.col("Covered (Secured)") / pl.col("Total EAD") * 100)
            .round(2)
            .alias("Security Coverage %")
        )
    return result


# ---------------------------------------------------------------------------
# 10. Write-off summary (all months)
# ---------------------------------------------------------------------------


def generate_writeoff_summary(df: pl.DataFrame) -> pl.DataFrame:
    group_cols = [c for c in ["written_off", "stage"] if c in df.columns]
    if not group_cols:
        return pl.DataFrame({"note": ["written_off / stage columns not available"]})

    agg: list[pl.Expr] = []
    if "ead" in df.columns:
        agg.append(pl.col("ead").cast(pl.Float64, strict=False).sum().alias("Total EAD"))
    if "total_provision" in df.columns:
        agg.append(
            pl.col("total_provision").cast(pl.Float64, strict=False).sum().alias("Total Provision")
        )
    if "loan_id" in df.columns:
        agg.append(pl.col("loan_id").count().alias("Loan Count"))
    if not agg:
        agg = [pl.len().alias("Loan Count")]

    return df.group_by(group_cols).agg(agg).sort(group_cols)


# ---------------------------------------------------------------------------
# 11. SBU / SubSBU breakdown (all months)
# ---------------------------------------------------------------------------


def generate_sbu_summary(df: pl.DataFrame) -> pl.DataFrame:
    group_cols = [c for c in ["sbu", "sub_sbu"] if c in df.columns]
    if not group_cols:
        return pl.DataFrame({"note": ["sbu / sub_sbu columns not available"]})

    agg: list[pl.Expr] = []
    if "ead" in df.columns:
        agg.append(pl.col("ead").cast(pl.Float64, strict=False).sum().alias("Total EAD"))
    if "total_provision" in df.columns:
        agg.append(
            pl.col("total_provision").cast(pl.Float64, strict=False).sum().alias("Total Provision")
        )
    if "loan_id" in df.columns:
        agg.append(pl.col("loan_id").count().alias("Loan Count"))
    if not agg:
        agg = [pl.len().alias("Loan Count")]

    rename = {"sbu": "SBU", "sub_sbu": "Sub-SBU"}
    sort_col = rename.get(group_cols[0], group_cols[0])
    return (
        df.group_by(group_cols)
        .agg(agg)
        .rename({k: v for k, v in rename.items() if k in group_cols})
        .sort(sort_col)
    )


# ---------------------------------------------------------------------------
# 12. Provision reasonableness check (all months)
# ---------------------------------------------------------------------------


def generate_provision_check(df: pl.DataFrame) -> pl.DataFrame:
    if "additional_provision" not in df.columns:
        return pl.DataFrame({"note": ["additional_provision column not available"]})

    ap = df["additional_provision"].cast(pl.Float64, strict=False).fill_null(0.0)
    filtered = df.filter(ap > 0)

    if filtered.is_empty():
        return pl.DataFrame({"note": ["No rows with additional_provision > 0"]})

    select_cols = [c for c in ["loan_id", "stage", "dpd_bucket", "ead"] if c in filtered.columns]
    result = filtered.select(select_cols + ["additional_provision"])

    if "existing_provision" in filtered.columns:
        result = result.with_columns(
            filtered["existing_provision"]
            .cast(pl.Float64, strict=False)
            .alias("existing_provision")
        )
        result = result.with_columns(
            pl.when(pl.col("existing_provision") > 0)
            .then(
                pl.col("additional_provision").cast(pl.Float64, strict=False)
                / pl.col("existing_provision")
            )
            .otherwise(None)
            .round(4)
            .alias("Ratio (Addl/Policy)")
        )
        ratios = result["Ratio (Addl/Policy)"].drop_nulls()
        if ratios.len() > 0:
            median_ratio = ratios.median()
            result = result.with_columns(
                (pl.col("Ratio (Addl/Policy)") > 2 * median_ratio).alias("Outlier Flag")
            )

    rename = {
        "loan_id": "Loan ID",
        "stage": "Stage",
        "dpd_bucket": "DPD Bucket",
        "ead": "EAD",
        "additional_provision": "Additional Provision",
        "existing_provision": "Policy Provision",
    }
    return result.rename({k: v for k, v in rename.items() if k in result.columns})


# ---------------------------------------------------------------------------
# 13. Negative values check (all months)
# ---------------------------------------------------------------------------


def generate_negative_check(df: pl.DataFrame) -> pl.DataFrame:
    numeric_canonical = [
        "ead",
        "outstanding_principal",
        "outstanding_interest",
        "accrued_interest",
        "overdue_charges",
        "future_pos",
        "gross_book_value",
        "collateral_value",
        "covered_portion",
        "uncovered_portion",
        "existing_provision",
        "additional_provision",
        "total_provision",
        "sanction_amount",
        "disbursed_amount",
        "principal_paid",
        "interest_paid",
        "emi_amount",
        "mob",
    ]
    numeric_raw = [
        "Prepaid",
        "TotalDebtors",
        "Debtors_Redemption_Premium",
        "AccruedRedemptionPremium",
        "zero_90_days_interest",
        "zero_90_int_Final",
        "PROCESSINGFEES",
        "Unbanked",
    ]
    all_numeric = [c for c in (numeric_canonical + numeric_raw) if c in df.columns]

    rows = []
    total = len(df)
    for col in all_numeric:
        series = df[col].cast(pl.Float64, strict=False)
        neg_count = (series < 0).sum()
        null_count = series.is_null().sum()
        rows.append(
            {
                "Column": col,
                "Total Rows": total,
                "Negative Count": neg_count,
                "Null Count": null_count,
                "Has Issue": neg_count > 0,
            }
        )

    return pl.DataFrame(rows) if rows else pl.DataFrame({"note": ["No numeric columns found"]})


# ---------------------------------------------------------------------------
# 14. Stage-DPD mismatch (all months)
# ---------------------------------------------------------------------------


def generate_stage_mismatch(df: pl.DataFrame) -> pl.DataFrame:
    if "stage" not in df.columns:
        return pl.DataFrame({"note": ["stage column not available"]})

    stage_str = df["stage"].cast(pl.Utf8, strict=False).str.strip_chars()
    conditions = []

    if "dpd_bucket" in df.columns:
        high_dpd = _is_high_dpd(df["dpd_bucket"])
        conditions.append(high_dpd & (stage_str == "1"))

    if "loan_status" in df.columns:
        ls_lower = df["loan_status"].cast(pl.Utf8, strict=False).str.to_lowercase().fill_null("")
        is_closed = ls_lower.map_elements(
            lambda v: any(s in v for s in _CLOSED_STATUSES),
            return_dtype=pl.Boolean,
        )
        stage_23 = stage_str.is_in(["2", "3"])
        conditions.append(is_closed & stage_23)

    if not conditions:
        return pl.DataFrame({"note": ["Insufficient columns to detect mismatches"]})

    mask = conditions[0]
    for c in conditions[1:]:
        mask = mask | c

    result = df.filter(mask)
    if result.is_empty():
        return pl.DataFrame({"note": ["No stage-DPD mismatches found — data is consistent"]})

    keep_cols = [
        c for c in ["loan_id", "stage", "dpd_bucket", "loan_status", "ead"] if c in result.columns
    ]
    return result.select(keep_cols).rename(
        {
            k: v
            for k, v in {
                "loan_id": "Loan ID",
                "stage": "Stage",
                "dpd_bucket": "DPD Bucket",
                "loan_status": "Loan Status",
                "ead": "EAD",
            }.items()
            if k in keep_cols
        }
    )


# ---------------------------------------------------------------------------
# 15. FVTPL split (all months)
# ---------------------------------------------------------------------------


def generate_fvtpl_split(df: pl.DataFrame) -> pl.DataFrame:
    if "FVTPL" not in df.columns:
        return pl.DataFrame(
            {"note": ["FVTPL column not available (not mapped or not in source data)"]}
        )

    agg: list[pl.Expr] = []
    if "ead" in df.columns:
        agg.append(pl.col("ead").cast(pl.Float64, strict=False).sum().alias("Total EAD"))
    if "total_provision" in df.columns:
        agg.append(
            pl.col("total_provision").cast(pl.Float64, strict=False).sum().alias("Total Provision")
        )
    if "loan_id" in df.columns:
        agg.append(pl.col("loan_id").count().alias("Loan Count"))
    if not agg:
        agg = [pl.len().alias("Loan Count")]

    return df.group_by("FVTPL").agg(agg).sort("FVTPL")


# ---------------------------------------------------------------------------
# Report registry — order matters (primary 5 first, additional 10 after)
# ---------------------------------------------------------------------------

_REPORT_FUNCTIONS: list[tuple[str, object, str]] = [
    ("ucid_lan", generate_ucid_lan_count, "UCID → LAN Count"),
    ("state_ead", generate_state_ead, "State-wise EAD (Last Month)"),
    ("disbursement", generate_disbursement_summary, "Disbursements in Audit Period"),
    ("product_ead", generate_product_ead, "Product-wise EAD (Last Month)"),
    ("product_stage_dpd", generate_product_stage_dpd, "Product × Stage × DPD"),
    ("pivot", generate_pivot_report, "Master Pivot Report"),
    ("stage_summary", generate_stage_summary, "Stage Summary"),
    ("dpd_summary", generate_dpd_summary, "DPD Bucket Summary"),
    ("security", generate_security_summary, "Secured vs Unsecured"),
    ("writeoff", generate_writeoff_summary, "Write-Off Summary"),
    ("sbu", generate_sbu_summary, "SBU - SubSBU Breakdown"),
    ("provision_check", generate_provision_check, "Provision Reasonableness"),
    ("negative_check", generate_negative_check, "Negative Values Check"),
    ("stage_mismatch", generate_stage_mismatch, "Stage-DPD Mismatch"),
    ("fvtpl", generate_fvtpl_split, "FVTPL Split"),
    # EAD summary reports (re-enabled; see fix B in v0.1.38 hardening)
    ("portfolio_concentration", generate_portfolio_concentration, "Portfolio Concentration"),
    ("stage_distribution", generate_stage_distribution, "Stage Distribution"),
    ("dpd_risk_distribution", generate_dpd_risk_distribution, "DPD-Risk Distribution"),
    ("collateral_coverage", generate_collateral_coverage, "Collateral Coverage"),
    ("provision_coverage", generate_provision_coverage, "Provision Coverage"),
    ("writeoff_recovery", generate_writeoff_recovery, "Write-off & Recovery"),
    ("sanction_disbursement", generate_sanction_disbursement_variance, "Sanction vs Disbursement"),
    ("data_quality_ead", generate_data_quality_summary_ead, "Data Quality Summary"),
]

# Derived list of (key, label) used by the API and templates
_REPORT_SHEET_NAMES: list[tuple[str, str]] = [(k, lbl) for k, _, lbl in _REPORT_FUNCTIONS]


# ---------------------------------------------------------------------------
# Excel workpaper builder
# ---------------------------------------------------------------------------


def _is_monetary_col(col: str) -> bool:
    """Return True if column likely contains monetary amounts (not counts/IDs/dates)."""
    c = col.lower()
    if c.endswith(("_count", "_pct", "_id", "_name", "_code", "_bucket",
                   "_flag", "_date", "_month", "_desc", "_system")):
        return False
    if c in {"count", "stage", "ucid", "lan_count", "new_loans",
              "dpd_stage_mismatch", "score", "ratio"}:
        return False
    return any(p in c for p in (
        "ead", "provision", "outstanding", "principal", "disburse",
        "sanction", "writeoff", "write_off", "recovery", "amount",
        "balance", "interest", "premium", "exposure",
    ))


def _write_sheet(
    ws,
    df: pl.DataFrame,
    header_fill,
    header_font,
    data_font,
    *,
    monetary_cols: frozenset[str] = frozenset(),
) -> None:
    """Write a Polars DataFrame to an openpyxl worksheet.

    Preserves numeric types so Excel can format them (₹, %, etc).
    Monetary columns are written as formulas dividing by Settings!$B$4 so the
    unit-converter dropdown on the Settings sheet scales them live.
    """
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    cols = df.columns
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    _numeric_dtypes = (
        pl.Int8, pl.Int16, pl.Int32, pl.Int64,
        pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64,
        pl.Float32, pl.Float64,
    )
    # Identify which column indices are numeric (0-based)
    numeric_col_idx: set[int] = {
        i for i, col in enumerate(cols) if df[col].dtype in _numeric_dtypes
    }
    # Identify which numeric columns are monetary (need divisor formula)
    monetary_col_idx: set[int] = {
        i for i in numeric_col_idx if cols[i] in monetary_cols
    }

    # Cast non-numeric to string for row iteration
    casts = [
        pl.col(col).cast(pl.Utf8, strict=False)
        for col in cols
        if df[col].dtype not in _numeric_dtypes
    ]
    df_for_write = df.with_columns(casts) if casts else df

    for ri, row in enumerate(df_for_write.rows(), 2):
        for ci, val in enumerate(row, 1):
            if val in (None, "None", "null"):
                ws.cell(row=ri, column=ci, value=None).font = data_font
            elif (ci - 1) in monetary_col_idx and val is not None:
                # Monetary cell — formula divides by Settings!$B$4 (the unit divisor)
                cell = ws.cell(row=ri, column=ci, value=f"={val}/Settings!$B$4")
                cell.font = data_font
            else:
                ws.cell(row=ri, column=ci, value=val).font = data_font

    for ci, col in enumerate(cols, 1):
        max_len = min(max(len(col), 10), 40)
        ws.column_dimensions[get_column_letter(ci)].width = max_len + 2

    if len(df) > 0:
        ws.freeze_panes = "A2"


def _build_settings_sheet(ws) -> None:
    """Write the unit-converter Settings sheet (always the first sheet)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    ACCENT = "5C3D1E"   # terracotta (matches house style)
    BG = "FDF6EE"       # warm beige

    ws.title = "Settings"
    ws.sheet_properties.tabColor = ACCENT

    # Header banner
    ws["A1"] = "EAD Workpaper — Display Settings"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:C1")

    # Unit selector row
    ws["A3"] = "Display Unit"
    ws["A3"].font = Font(bold=True, size=11)
    ws["B3"] = "Lakhs (₹ Lakhs)"   # default
    ws["B3"].font = Font(size=11, bold=True, color=ACCENT)
    ws["B3"].fill = PatternFill(start_color=BG, end_color=BG, fill_type="solid")
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

    # Data validation dropdown
    dv = DataValidation(
        type="list",
        formula1='"Actual (₹),Lakhs (₹ Lakhs),Crores (₹ Cr),Millions (₹ Mn)"',
        allow_blank=False,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add("B3")

    # Divisor formula (evaluated by Excel, drives all monetary cells)
    ws["A4"] = "Divisor"
    ws["A4"].font = Font(size=10, color="888888")
    ws["B4"] = (
        '=IF(B3="Actual (₹)",1,'
        'IF(B3="Lakhs (₹ Lakhs)",100000,'
        'IF(B3="Crores (₹ Cr)",10000000,1000000)))'
    )
    ws["B4"].font = Font(size=10, color="888888")

    # Unit label formula (for display)
    ws["A5"] = "Unit Label"
    ws["A5"].font = Font(size=10, color="888888")
    ws["B5"] = (
        '=IF(B3="Actual (₹)","₹",'
        'IF(B3="Lakhs (₹ Lakhs)","₹ Lakhs",'
        'IF(B3="Crores (₹ Cr)","₹ Cr","₹ Mn")))'
    )
    ws["B5"].font = Font(size=10, color="888888")

    # Instructions
    ws["A7"] = "How to use"
    ws["A7"].font = Font(bold=True, size=10)
    ws["B7"] = (
        "Click the dropdown in cell B3 and select your preferred unit. "
        "All monetary values across every sheet (EAD, Provision, Outstanding, etc.) "
        "update automatically. Counts and percentages are unaffected."
    )
    ws["B7"].font = Font(size=10)
    ws["B7"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[7].height = 45

    ws["A9"] = "Lead File Note"
    ws["A9"].font = Font(bold=True, size=10)
    ws["B9"] = (
        "If you link this workpaper to a lead file, the lead file values "
        "will also update when you change the unit here — all formulas divide by Settings!$B$4."
    )
    ws["B9"].font = Font(size=10)
    ws["B9"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[9].height = 45

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 8


def build_ead_workpaper(reports: dict[str, pl.DataFrame], output_path: Path) -> None:
    """Write all ran EAD reports into one multi-sheet Excel workpaper.

    Sheet 1 is always "Settings" — a unit-converter dropdown (Actual / Lakhs / Crores /
    Millions) whose divisor value is referenced by every monetary cell across all data sheets.
    Changing the dropdown instantly rescales all amounts; counts and percentages are unaffected.
    """
    import openpyxl
    from openpyxl.styles import Font

    from fcmr_core.reporting.excel_style import HEADER_FILL, HEADER_FONT

    wb = openpyxl.Workbook()

    # Sheet 1: Settings (unit converter)
    ws_settings = wb.active
    _build_settings_sheet(ws_settings)

    header_fill = HEADER_FILL
    header_font = HEADER_FONT
    data_font = Font(size=10)

    has_data = False
    for key, sheet_name in _REPORT_SHEET_NAMES:
        df = reports.get(key)
        if df is None or df.is_empty():
            continue
        ws = wb.create_sheet(sheet_name[:31])
        monetary_cols = frozenset(col for col in df.columns if _is_monetary_col(col))
        _write_sheet(ws, df, header_fill, header_font, data_font, monetary_cols=monetary_cols)
        has_data = True

    if not has_data:
        ws = wb.create_sheet("Summary")
        ws["A1"] = "No data available"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# Top-level runner
# ---------------------------------------------------------------------------


def run_ead_analytics(
    df: pl.DataFrame,
    output_dir: Path,
    report_keys: list[str] | None = None,
    period_from: str | None = None,
    period_to: str | None = None,
    on_progress=None,
) -> dict[str, Path]:
    """Run EAD analytics reports, write CSVs + Excel workpaper, return {key: path}.

    Args:
        df: Merged EAD DataFrame (all uploads for the engagement).
        output_dir: Directory to write CSV and Excel output files.
        report_keys: List of report keys to run. ``None`` runs all 15.
        period_from: Engagement period start date (ISO string, e.g. "2026-04-01").
            Used by the disbursement analytics to filter rows.
        period_to: Engagement period end date (ISO string).
        on_progress: Optional callback(completed, total, label) for progress tracking.
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    # Filter to requested reports; None = all
    to_run: list[tuple[str, object, str]] = (
        [(k, fn, lbl) for k, fn, lbl in _REPORT_FUNCTIONS if k in report_keys]
        if report_keys is not None
        else list(_REPORT_FUNCTIONS)
    )

    total = len(to_run) + 1  # +1 for workpaper step
    reports: dict[str, pl.DataFrame] = {}
    paths: dict[str, Path] = {}

    for i, (key, fn, label) in enumerate(to_run):
        if on_progress:
            on_progress(i, total, label)
        try:
            # Disbursement function needs the engagement period dates
            if key == "disbursement":
                result = fn(df, period_from=period_from, period_to=period_to)  # type: ignore[call-arg]
            else:
                result = fn(df)  # type: ignore[call-arg]
            reports[key] = result
            if not result.is_empty():
                csv_path = output_dir / f"{key}.csv"
                result.write_csv(str(csv_path))
                paths[key] = csv_path
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"EAD report '{key}' failed: {exc}")
            reports[key] = pl.DataFrame({"note": [f"Error: {exc}"]})

    if on_progress:
        on_progress(len(to_run), total, "Building Excel workpaper")

    workpaper_path = output_dir / "ead_workpaper.xlsx"
    try:
        build_ead_workpaper(reports, workpaper_path)
        paths["workpaper"] = workpaper_path
    except Exception:
        pass

    if on_progress:
        on_progress(total, total, "Complete")

    return paths


# ---------------------------------------------------------------------------
# Summary stats (for the result page header cards)
# ---------------------------------------------------------------------------


def compute_summary_stats(df: pl.DataFrame) -> dict:
    """Return high-level numbers for the four stat cards on the results page."""
    stats: dict = {}

    if "loan_id" in df.columns:
        stats["total_loans"] = df["loan_id"].count()
    elif len(df) > 0:
        stats["total_loans"] = len(df)
    else:
        stats["total_loans"] = 0

    if "ead" in df.columns:
        stats["total_ead"] = df["ead"].cast(pl.Float64, strict=False).sum()
    else:
        stats["total_ead"] = None

    if "total_provision" in df.columns:
        stats["total_provision"] = df["total_provision"].cast(pl.Float64, strict=False).sum()
    else:
        stats["total_provision"] = None

    if stats.get("total_ead") and stats.get("total_provision") and stats["total_ead"] > 0:
        stats["coverage_pct"] = round(stats["total_provision"] / stats["total_ead"] * 100, 2)
    else:
        stats["coverage_pct"] = None

    return stats
