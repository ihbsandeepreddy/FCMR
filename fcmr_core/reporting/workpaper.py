"""Excel workpaper builder for audit documentation.

Generates a 5-sheet workpaper: Cover, Lead Sheet, Detailed Exceptions, TOC and TOD, Methodology.
"""

from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fcmr_core.analytics.cm_summary import (
    generate_cluster_distribution,
    generate_coapplicant_overlap,
    generate_data_quality_summary,
    generate_demographic_distribution,
    generate_duplication_summary,
    generate_geographic_distribution,
    generate_kyc_completeness,
    generate_lan_concentration,
)
from fcmr_core.catalog import store
from fcmr_core.logging_setup import get_logger
from fcmr_core.reporting.excel_style import (
    HEADER_FILL,
    HEADER_FONT,
    QUANTITY_FORMAT_INT,
    THIN_BORDER,
    add_signoff_block,
    apply_header_style,
    auto_column_widths,
    freeze_header,
)
from fcmr_core.rules.registry import CATEGORIES, list_rules
from fcmr_core.sampling.stratification import _SEVERITY_MAP


def _sanitize_filename(text: str) -> str:
    """Remove invalid filename characters."""
    invalid = '<>:"/\\|?*'
    for char in invalid:
        text = text.replace(char, "_")
    return text


def _procedures_performed(run: dict, long_csv_path: Path) -> list[dict]:
    """Return list of procedures (rules) that were run with metadata.

    Returns: [{"rule_id": str, "description": str, "category": str,
               "severity": str, "exceptions": int}, ...]
    """
    try:
        # Determine which rules were run
        selected_rules_json = run.get("selected_rules")
        if selected_rules_json:
            selected_rule_ids = json.loads(selected_rules_json)
            rule_universe = [r for r in list_rules() if r.rule_id in selected_rule_ids]
        else:
            rule_universe = list(list_rules())

        # Build category map
        category_map = {}
        for cat in CATEGORIES:
            for rule_id in cat["rule_ids"]:
                category_map[rule_id] = cat["label"]

        # Read long CSV once; derive counts + severities from single pass
        try:
            long_df = pl.read_csv(long_csv_path, infer_schema_length=0)

            # Exception counts per rule (filter to non-OK, group, count)
            exception_counts = (
                long_df.filter(pl.col("status") != "OK").group_by("rule_id").len().to_dicts()
            )
            exc_map = {d["rule_id"]: d["len"] for d in exception_counts}

            # Severity per rule: collect unique codes per rule, map to severity
            rule_codes_agg = (
                long_df.group_by("rule_id").agg(pl.col("exception_code").unique()).to_dicts()
            )
            severity_map = {}
            sev_order = {"CRITICAL": 3, "HIGH": 2, "MEDIUM": 1, "LOW": 0}
            for item in rule_codes_agg:
                rule_id = item["rule_id"]
                codes = item["exception_code"]
                if codes:
                    severities = [_SEVERITY_MAP.get(code, "LOW") for code in codes]
                    severity = max(severities, key=lambda s: sev_order.get(s, 0))
                else:
                    severity = "—"
                severity_map[rule_id] = severity
        except Exception:
            exc_map = {}
            severity_map = {}

        # Build procedures list
        procedures = []
        for rule_meta in rule_universe:
            rule_id = rule_meta.rule_id
            category = category_map.get(rule_id, "—")
            exceptions = exc_map.get(rule_id, 0)
            severity = severity_map.get(rule_id, "—")

            procedures.append(
                {
                    "rule_id": rule_id,
                    "description": rule_meta.description,
                    "category": category,
                    "severity": severity,
                    "exceptions": exceptions,
                }
            )

        return procedures
    except Exception:
        return []


def build_workpaper(
    engagement: dict,
    run: dict,
    upload: dict,
    wide_csv_path: Path,
    long_csv_path: Path,
    sample_records: list[dict],
    output_dir: Path,
) -> Path:
    """Build a 13-sheet Excel workpaper (Cover, Lead, Detailed, TOC/TOD, Methodology, + 8 CM Summaries).

    Args:
        engagement: Engagement dict from store.
        run: Run dict from store.
        upload: Upload dict from store (filename, row_count, ingested_at, etc.).
        wide_csv_path: Path to wide exception CSV.
        long_csv_path: Path to long exception CSV.
        sample_records: List of sampled records from select_sample().
        output_dir: Directory to save the workpaper.

    Returns:
        Path to saved workpaper .xlsx.
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    # Generate filename
    engagement_name = _sanitize_filename(engagement.get("name", "Engagement"))
    period_from = (engagement.get("period_from") or "").split("T")[0]
    period_to = (engagement.get("period_to") or "").split("T")[0]
    timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    filename = f"{engagement_name}_{period_from}_{period_to}_{timestamp}.xlsx"
    workpaper_path = output_dir / filename

    # Create workbook
    wb = Workbook()
    ws_cover = wb.active
    ws_cover.title = "Cover"

    # Styles (use shared utilities)
    header_fill = HEADER_FILL
    header_font = HEADER_FONT
    title_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    border = THIN_BORDER

    # ── Sheet 0: Cover ──
    _build_cover_sheet(
        ws_cover,
        run,
        engagement,
        wide_csv_path,
        title_font,
        subheader_font,
        header_fill,
        header_font,
    )

    # ── Sheet 1: Lead Sheet ──
    ws1 = wb.create_sheet("Lead Sheet")
    _build_lead_sheet(
        ws1,
        engagement,
        run,
        upload,
        wide_csv_path,
        long_csv_path,
        header_fill,
        header_font,
        title_font,
        subheader_font,
        border,
    )

    # ── Sheet 2: Detailed Exceptions ──
    ws2 = wb.create_sheet("Detailed Exceptions")
    _build_detailed_exceptions_sheet(
        ws2, wide_csv_path, header_fill, header_font, border, max_rows=50000
    )

    # ── Sheet 3: TOC and TOD ──
    ws3 = wb.create_sheet("TOC and TOD")
    _build_toc_tod_sheet(ws3, sample_records, header_fill, header_font, border)

    # ── Sheet 4: Methodology ──
    ws4 = wb.create_sheet("Methodology")
    _build_methodology_sheet(
        ws4, engagement, run, wide_csv_path, sample_records, title_font, subheader_font
    )

    # ── Sheets 5-12: CM Summary Reports (if customer_master) ──
    if upload and upload.get("report_type") == "customer_master":
        try:
            df = store.get_upload_df(upload["upload_id"])
            if df is not None and not df.is_empty():
                # All 8 summaries
                summaries = [
                    ("Geographic Distribution", generate_geographic_distribution(df)),
                    ("KYC Completeness", generate_kyc_completeness(df)),
                    ("Demographic Distribution", generate_demographic_distribution(df)),
                    ("Duplication Summary", generate_duplication_summary(df)),
                    ("Co-Applicant Overlap", generate_coapplicant_overlap(df)),
                    ("Related-Party Clusters", generate_cluster_distribution(df)),
                    ("Data Quality Summary", generate_data_quality_summary(df)),
                    ("LAN Concentration", generate_lan_concentration(df)),
                ]
                for title, summary_df in summaries:
                    if (
                        summary_df
                        and not summary_df.is_empty()
                        and "note" not in summary_df.columns
                    ):
                        ws = wb.create_sheet(title)
                        _write_summary_sheet(
                            ws, title, summary_df, header_fill, header_font, border
                        )
        except Exception:
            pass  # Silently skip summaries if there's an error

    # Save
    wb.save(workpaper_path)
    return workpaper_path


def _build_cover_sheet(
    ws,
    run,
    engagement,
    wide_csv_path,
    title_font,
    subheader_font,
    header_fill,
    header_font,
):
    """Build Cover sheet."""
    row = 1
    report_type = run.get("upload_id", "—")[:20]

    # Title
    ws[f"A{row}"] = "SanGir Automations — Audit Working Paper"
    ws[f"A{row}"].font = title_font
    row += 2

    # W/P Reference and metadata
    ws[f"A{row}"] = "W/P Reference:"
    ws[f"B{row}"] = f"WP-{report_type}-{run.get('run_id', 'N/A')[:8]}"
    row += 1

    ws[f"A{row}"] = "Engagement:"
    ws[f"B{row}"] = engagement.get("name", "N/A")
    row += 1

    ws[f"A{row}"] = "Client:"
    ws[f"B{row}"] = engagement.get("client_name", "N/A")
    row += 1

    ws[f"A{row}"] = "Period:"
    period_str = f"{engagement.get('period_from', '')} to {engagement.get('period_to', '')}"
    ws[f"B{row}"] = period_str
    row += 1

    ws[f"A{row}"] = "Report Type:"
    ws[f"B{row}"] = report_type
    row += 1

    try:
        pop = len(pl.read_csv(wide_csv_path, infer_schema_length=0))
        ws[f"A{row}"] = "Population (records):"
        ws[f"B{row}"] = pop
    except Exception:
        pass
    row += 1

    ws[f"A{row}"] = "Date Prepared:"
    ws[f"B{row}"] = datetime.now(UTC).strftime("%Y-%m-%d")
    row += 2

    # Sign-off block (prepared by)
    add_signoff_block(ws, row, "Prepared By")
    row += 4

    ws[f"A{row}"] = "Date:"
    ws[f"B{row}"] = ""
    ws[f"C{row}"] = ""
    row += 2

    ws[f"A{row}"] = "Reviewed By:"
    ws[f"B{row}"] = ""
    ws[f"C{row}"] = ""
    row += 1

    ws[f"A{row}"] = "Signature:"
    ws[f"B{row}"] = ""
    ws[f"C{row}"] = ""
    row += 1

    ws[f"A{row}"] = "Date:"
    ws[f"B{row}"] = ""
    ws[f"C{row}"] = ""
    row += 2

    # Index of working papers
    ws[f"A{row}"] = "Index of Working Papers"
    ws[f"A{row}"].font = subheader_font
    row += 1

    headers = ["Sheet", "Purpose"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    row += 1

    sheets_info = [
        ("Lead Sheet", "Summary, procedures & conclusion"),
        ("Detailed Exceptions", "Full population with per-rule results"),
        ("TOC and TOD", "Sample testing & sign-off"),
        ("Methodology", "Sampling basis & ICFR control mapping"),
    ]
    for sheet_name, purpose in sheets_info:
        ws[f"A{row}"] = sheet_name
        ws[f"B{row}"] = purpose
        row += 1

    row += 1

    # Tickmark legend
    ws[f"A{row}"] = "Tickmark Legend"
    ws[f"A{row}"].font = subheader_font
    row += 1

    tickmarks = [
        "✓ = Agreed to source document, no exception",
        "Ø = Exception noted",
        "N/A = Not applicable",
        "S = Selected for sample testing",
    ]
    for mark in tickmarks:
        ws[f"A{row}"] = mark
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18


def _build_lead_sheet(
    ws,
    engagement,
    run,
    upload,
    wide_csv_path,
    long_csv_path,
    header_fill,
    header_font,
    title_font,
    subheader_font,
    border,
):
    """Build Lead Sheet."""
    row = 1

    # Header
    ws[f"A{row}"] = "SanGir Automations - Audit Workpaper"
    ws[f"A{row}"].font = title_font
    row += 1

    ws[f"A{row}"] = f"Engagement: {engagement.get('name', 'N/A')}"
    ws[f"A{row}"].font = subheader_font
    row += 1

    ws[f"A{row}"] = (
        f"Client: {engagement.get('client_name', 'N/A')} | Period: {engagement.get('period_from', '')} to {engagement.get('period_to', '')}"
    )
    row += 1

    ws[f"A{row}"] = f"Audit Date: {datetime.now(UTC).strftime('%Y-%m-%d')}"
    row += 2

    # Purpose & Objective
    ws[f"A{row}"] = "Purpose & Objective"
    ws[f"A{row}"].font = subheader_font
    row += 1

    purpose_text = (
        "This engagement validates the customer_master KYC and data quality via deterministic rules "
        "aligned with RBI Know Your Customer (KYC) Guidelines, ICAI Audit Sampling Guidance, and NFRA fraud-risk indicators."
    )
    ws[f"A{row}"] = purpose_text
    ws[f"A{row}"].alignment = Alignment(wrap_text=True)
    row += 2

    # Population Reconciliation
    ws[f"A{row}"] = "Source of Data / Population Reconciliation"
    ws[f"A{row}"].font = subheader_font
    row += 1

    reconciliation_headers = [
        "Source File",
        "Rows Ingested",
        "Records Analyzed",
        "Difference",
        "Ingested At",
    ]
    for col_idx, header in enumerate(reconciliation_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    row += 1

    try:
        wide_df = pl.read_csv(wide_csv_path, infer_schema_length=0)
        analyzed = len(wide_df)
        ingested = upload.get("row_count", analyzed)
        difference = ingested - analyzed

        ws[f"A{row}"] = upload.get("filename", "—")
        ws[f"B{row}"] = ingested
        ws[f"C{row}"] = analyzed
        ws[f"D{row}"] = difference
        ws[f"E{row}"] = upload.get("ingested_at", "—")[:10] if upload.get("ingested_at") else "—"
        row += 2
    except Exception:
        row += 2

    # Exception Summary
    ws[f"A{row}"] = "Exception Summary"
    ws[f"A{row}"].font = subheader_font
    row += 1

    try:
        df = pl.read_csv(wide_csv_path, columns=["overall_status"], infer_schema_length=0)
        status_counts = df["overall_status"].value_counts().to_dicts()
        status_dict = {d["overall_status"]: d["count"] for d in status_counts}

        total = sum(status_dict.values())
        ws[f"A{row}"] = "OK:"
        ws[f"B{row}"] = status_dict.get("OK", 0)
        ws[f"C{row}"] = f"{(status_dict.get('OK', 0) / total * 100):.1f}%" if total > 0 else "0%"
        row += 1

        ws[f"A{row}"] = "Warnings:"
        ws[f"B{row}"] = status_dict.get("WARN", 0)
        ws[f"C{row}"] = f"{(status_dict.get('WARN', 0) / total * 100):.1f}%" if total > 0 else "0%"
        row += 1

        ws[f"A{row}"] = "Errors:"
        ws[f"B{row}"] = status_dict.get("ERROR", 0)
        ws[f"C{row}"] = f"{(status_dict.get('ERROR', 0) / total * 100):.1f}%" if total > 0 else "0%"
        row += 2
    except Exception:
        row += 4

    # Procedures Performed
    ws[f"A{row}"] = "Procedures Performed"
    ws[f"A{row}"].font = subheader_font
    row += 1

    proc_headers = [
        "#",
        "Rule ID",
        "Audit Procedure (Description)",
        "Category",
        "Severity",
        "Exceptions",
        "Exception %",
    ]
    for col_idx, header in enumerate(proc_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    row += 1

    try:
        procedures = _procedures_performed(run, long_csv_path)
        total_records = len(pl.read_csv(wide_csv_path, infer_schema_length=0))
        for idx, proc in enumerate(procedures, 1):
            ws[f"A{row}"] = idx
            ws[f"B{row}"] = proc["rule_id"]
            ws[f"C{row}"] = proc["description"]
            ws[f"D{row}"] = proc["category"]
            ws[f"E{row}"] = proc["severity"]
            ws[f"F{row}"] = proc["exceptions"]
            ws[f"F{row}"].number_format = QUANTITY_FORMAT_INT
            exc_pct = (proc["exceptions"] / total_records * 100) if total_records > 0 else 0
            ws[f"G{row}"] = f"{exc_pct:.1f}%"
            row += 1
    except Exception:
        pass

    row += 1

    # Results & Conclusion
    ws[f"A{row}"] = "Results & Conclusion"
    ws[f"A{row}"].font = subheader_font
    row += 1

    try:
        df = pl.read_csv(wide_csv_path, infer_schema_length=0)
        total_recs = len(df)
        exc_recs = sum(1 for val in df["overall_status"] if val != "OK")
        exc_rate = (exc_recs / total_recs * 100) if total_recs > 0 else 0

        conclusion = (
            f"Analysis of {total_recs:,} customer records identified {exc_recs:,} records ({exc_rate:.1f}%) "
            f"with one or more exceptions. All exceptions have been documented and ranked by severity. "
            f"Further investigation is recommended for CRITICAL and HIGH severity findings."
        )
        ws[f"A{row}"] = conclusion
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
    except Exception:
        ws[f"A{row}"] = "Results pending."

    # Column widths for Lead Sheet
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    # Note: no freeze/autofilter here — the Lead sheet is a non-tabular summary
    # (title, info, breakdown, procedures sub-table, prose), not a single table.

    # Add reviewed-by sign-off block at the end
    max_row = ws.max_row
    add_signoff_block(ws, max_row + 3, "Reviewed By")


def _build_detailed_exceptions_sheet(
    ws, wide_csv_path, header_fill, header_font, border, max_rows: int = 50000
):
    """Build Detailed Exceptions sheet (exception rows only, capped for scale safety).

    Args:
        ws: openpyxl worksheet
        wide_csv_path: Path to wide CSV
        header_fill: Header cell fill style
        header_font: Header cell font style
        border: Border style (currently unused but available)
        max_rows: Maximum rows to write (default 50,000); prevents OOM on large files
    """
    try:
        df = pl.read_csv(wide_csv_path, infer_schema_length=0)

        # Filter to exception rows only (overall_status != "OK")
        exc_df = df.filter(pl.col("overall_status") != "OK")

        # Log truncation if capping
        total_exceptions = len(exc_df)
        if total_exceptions > max_rows:
            get_logger("reporting").warning(
                "detailed_exceptions_truncated total=%d capped=%d",
                total_exceptions,
                max_rows,
            )
            exc_df = exc_df.head(max_rows)

        # Mask Aadhaar (vectorized): replace with XXXXXXXX + last 4 chars
        for col_name in exc_df.columns:
            if "aadhaar" in col_name.lower() or "aadhar" in col_name.lower():
                exc_df = exc_df.with_columns(
                    pl.when(pl.col(col_name).str.len_chars() >= 4)
                    .then(pl.lit("XXXXXXXX") + pl.col(col_name).str.slice(-4))
                    .otherwise(pl.col(col_name))
                    .alias(col_name)
                )

        col_names = exc_df.columns

        # Header row: "Ref #" + original columns
        ws.cell(row=1, column=1, value="Ref #").fill = header_fill
        ws.cell(row=1, column=1).font = header_font
        for col_idx, col_name in enumerate(col_names, 2):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font

        # Data rows: prepend ref number, then original columns
        for ref_num, row_vals in enumerate(exc_df.iter_rows(allow_na=True), 1):
            ws.cell(row=ref_num + 1, column=1, value=ref_num)
            for col_idx, val in enumerate(row_vals, 2):
                ws.cell(row=ref_num + 1, column=col_idx, value=val)

        # Freeze top row and enable autofilter
        freeze_header(ws)

        # Adjust column widths (cap at 18)
        auto_column_widths(ws, max_width=18)

        # Metadata row (below the data, before signing off)
        if total_exceptions > max_rows:
            metadata_row = len(exc_df) + 2
            ws[f"A{metadata_row}"] = (
                f"Note: Showing {max_rows:,} of {total_exceptions:,} exception rows. Full data available in exception CSVs."
            )
    except Exception:
        ws["A1"] = "Unable to load exception data"


def _build_toc_tod_sheet(ws, sample_records, header_fill, header_font, border):
    """Build Test of Controls / Test of Details sheet with ICFR attributes."""

    # ICFR attribute map based on criticality/exception codes
    def get_control_objective(criticality):
        if criticality == "CRITICAL":
            return "No duplicate / fraudulent customers"
        elif criticality == "HIGH":
            return "KYC data integrity"
        else:
            return "Data quality controls"

    def get_assertion(codes):
        # Simple mapping: if duplicates → Existence, if missing → Completeness, else Accuracy
        if "DUP" in codes or "DUPLICATE" in codes:
            return "Existence"
        elif "MISSING" in codes or "INCOMPLETE" in codes:
            return "Completeness"
        else:
            return "Accuracy"

    def get_attribute_tested(criticality, codes):
        if "DUP" in codes or "DUPLICATE" in codes:
            return "Duplicate checking"
        elif criticality == "CRITICAL":
            return "Identity fraud indicators"
        else:
            return "Data format & completeness"

    # Headers
    headers = [
        "Sample#",
        "Row_Index",
        "Criticality",
        "Selection_Reason",
        "Control_Objective",
        "Assertion",
        "Attribute_Tested",
        "Tested_By",
        "Date",
        "Sign_Off",
        "Notes",
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    apply_header_style(ws, row=1)

    # Apply borders to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border

    # Sample rows
    for sample_idx, sample in enumerate(sample_records, 2):
        ws[f"A{sample_idx}"] = sample_idx - 1
        ws[f"B{sample_idx}"] = sample["row_index"]
        ws[f"C{sample_idx}"] = sample["criticality"]
        ws[f"D{sample_idx}"] = sample["selection_reason"]
        ws[f"E{sample_idx}"] = get_control_objective(sample["criticality"])
        ws[f"F{sample_idx}"] = get_assertion(sample.get("exception_codes", ""))
        ws[f"G{sample_idx}"] = get_attribute_tested(
            sample["criticality"], sample.get("exception_codes", "")
        )
        # H, I, J left blank for tester sign-off

    # Adjust column widths
    for col_idx, width in enumerate([8, 10, 12, 28, 28, 14, 22, 15, 12, 12, 16], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _get_selected_category_ids(run: dict) -> set[str]:
    """Return the set of category ids that were selected in the run.

    If selected_rules is None/empty (run all), return all category ids.
    Otherwise, return the ids of categories with at least one selected rule.
    Keyed on CATEGORIES ids so the ICFR table can't drift from the registry.
    """
    all_ids = {cat["id"] for cat in CATEGORIES}

    selected_rules_json = run.get("selected_rules")
    if not selected_rules_json:
        # Run all rules → include all categories
        return all_ids

    try:
        selected_rule_ids = set(json.loads(selected_rules_json))
    except Exception:
        return all_ids

    selected_cats = {
        cat["id"] for cat in CATEGORIES if any(rid in selected_rule_ids for rid in cat["rule_ids"])
    }

    return selected_cats if selected_cats else all_ids


def _build_methodology_sheet(
    ws, engagement, run, wide_csv_path, sample_records, title_font, subheader_font
):
    """Build Sampling Methodology Note sheet."""
    row = 1

    ws[f"A{row}"] = "Deterministic Risk-Based Sampling Methodology"
    ws[f"A{row}"].font = title_font
    row += 2

    ws[f"A{row}"] = "Approach"
    ws[f"A{row}"].font = subheader_font
    row += 1

    methodology_text = [
        "• Stratified by exception severity (CRITICAL > HIGH > MEDIUM > LOW)",
        "• ICAI-ICFR attribute sampling table (95% confidence, 5% tolerable deviation)",
        "• Seeded random selection for reproducibility across re-runs",
        "• Each sample tagged with selection reason and criticality level",
    ]

    for text in methodology_text:
        ws[f"A{row}"] = text
        row += 1

    row += 1
    ws[f"A{row}"] = "Strata & Severity Weights"
    ws[f"A{row}"].font = subheader_font
    row += 1

    strata_desc = [
        "CRITICAL: PAN, Aadhaar, UCID inconsistencies (identity fraud risk)",
        "HIGH: Voter ID, Address, Bank Account duplicates (fraud indicators)",
        "MEDIUM: Email domain, age range, account length (data quality issues)",
        "LOW: PIN/address mismatches (lower audit impact)",
    ]

    for text in strata_desc:
        ws[f"A{row}"] = text
        row += 1

    row += 1
    ws[f"A{row}"] = "Confidence & Precision"
    ws[f"A{row}"].font = subheader_font
    row += 1

    try:
        df = pl.read_csv(wide_csv_path, infer_schema_length=0)
        population = len(df)
        exception_count = sum(1 for val in df["overall_status"] if val != "OK")
        sample_size = len(sample_records)

        ws[f"A{row}"] = f"Population: {population:,} records"
        row += 1
        ws[f"A{row}"] = (
            f"Exceptions: {exception_count:,} records ({exception_count/population*100:.1f}%)"
        )
        row += 1
        ws[f"A{row}"] = f"Sample Size: {sample_size} (from ICAI table, 95% confidence)"
        row += 1
        ws[f"A{row}"] = f"Sample Rate: {sample_size/population*100:.2f}%"
        row += 2
    except Exception:
        row += 4

    ws[f"A{row}"] = "International Standards Alignment"
    ws[f"A{row}"].font = subheader_font
    row += 1

    standards = [
        "• ICAI Audit Sampling Guidance (India)",
        "• ISA 530 Audit Sampling (IAASB - International)",
        "• RBI Know Your Customer (KYC) Guidelines",
        "• NFRA fraud-risk indicators for NBFC audits",
        "• QRB recommendations for independent audits",
    ]

    for text in standards:
        ws[f"A{row}"] = text
        row += 1

    row += 1
    ws[f"A{row}"] = "Fraud-Risk Focus"
    ws[f"A{row}"].font = subheader_font
    row += 1

    fraud_focus = (
        "Loan account duplication, identity fraud, KYC weaknesses, undisclosed "
        "conflicts of interest, and cash-flow anomalies are weighted highest in "
        "this sampling design. The deterministic approach ensures reproducibility "
        "and defensibility in regulatory reviews."
    )
    ws[f"A{row}"] = fraud_focus
    ws[f"A{row}"].alignment = Alignment(wrap_text=True)

    row += 2
    ws[f"A{row}"] = "ICFR Control Mapping"
    ws[f"A{row}"].font = subheader_font
    row += 1

    # ICFR Control Mapping table
    icfr_headers = ["Category", "Control Objective", "Assertion", "Standard"]
    for col_idx, header in enumerate(icfr_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
    row += 1

    # Keyed by CATEGORIES id (single source of truth) → (display label, objective,
    # assertion, standard). Order follows the registry's CATEGORIES order.
    icfr_mappings = [
        (
            "missing_data",
            "Missing Data",
            "Mandatory KYC fields captured",
            "Completeness",
            "RBI KYC",
        ),
        (
            "kyc_format",
            "KYC & Document Format",
            "Customer identity is valid & verifiable",
            "Accuracy",
            "RBI KYC / ICAI",
        ),
        (
            "address_pin",
            "Address & PIN",
            "Address is complete & valid",
            "Completeness",
            "RBI KYC / ICAI",
        ),
        (
            "duplicates",
            "Duplicate Detection",
            "No duplicate / fictitious customers",
            "Existence",
            "NFRA fraud indicators",
        ),
        (
            "identity_grouping",
            "Identity Grouping (UCID + Beneficiary)",
            "Related parties identified",
            "Existence",
            "ICAI / NFRA",
        ),
    ]

    selected_cat_ids = _get_selected_category_ids(run)
    for cat_id, label, obj, assertion, standard in icfr_mappings:
        if cat_id in selected_cat_ids:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = obj
            ws[f"C{row}"] = assertion
            ws[f"D{row}"] = standard
            row += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22


def _write_summary_sheet(ws, title: str, df: pl.DataFrame, header_fill, header_font, border):
    """Write a summary report sheet (one of 8 CM summaries).

    Args:
        ws: openpyxl worksheet
        title: Sheet title (used for display)
        df: Polars DataFrame with the summary data
        header_fill: Header fill color
        header_font: Header font
        border: Border style
    """
    # Write header row
    columns = df.columns
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Write data rows
    for row_idx, row in enumerate(df.to_dicts(), 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            # Right-align numeric columns
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Freeze header and auto-size columns
    freeze_header(ws)
    auto_column_widths(ws)
