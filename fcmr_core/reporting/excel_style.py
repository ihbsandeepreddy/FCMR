"""Centralized Excel styling for audit workpapers (CM and EAD).

One house style across all sheets and both modules: frozen headers, consistent
fonts, number formats, borders, and column widths.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Colors & fills
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(size=10)

# ─────────────────────────────────────────────────────────────────────────────
# Borders
# ─────────────────────────────────────────────────────────────────────────────

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ─────────────────────────────────────────────────────────────────────────────
# Number formats
# ─────────────────────────────────────────────────────────────────────────────

# ₹ with thousands separator, 2 decimals
RUPEE_FORMAT = '"₹"#,##0.00'

# ₹ with thousands separator, no decimals (for whole amounts)
RUPEE_FORMAT_INT = '"₹"#,##0'

# Percentage with 2 decimals
PERCENT_FORMAT = '0.00"%"'

# Thousands separator, 2 decimals (for quantities)
QUANTITY_FORMAT = "#,##0.00"

# Thousands separator, no decimals
QUANTITY_FORMAT_INT = "#,##0"

# Date format (YYYY-MM-DD)
DATE_FORMAT = "YYYY-MM-DD"

# Count format (no decimals, thousands separator)
COUNT_FORMAT = "#,##0"


def apply_header_style(ws, row: int = 1):
    """Apply header styling to the first row of a worksheet.

    Args:
        ws: openpyxl worksheet
        row: Row number to style as header (default 1)
    """
    for col_idx, col_letter in enumerate(
        [get_column_letter(i) for i in range(1, ws.max_column + 1)], 1
    ):
        cell = ws[f"{col_letter}{row}"]
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def freeze_header(ws):
    """Freeze the first row and enable autofilter."""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def set_column_width(ws, col_idx: int, width: int | None = None):
    """Set a column width, capping at 40 for readability.

    Args:
        ws: openpyxl worksheet
        col_idx: Column index (1-based)
        width: Desired width; if None, auto-set to 20
    """
    w = width or 20
    ws.column_dimensions[get_column_letter(col_idx)].width = min(w, 40)


def auto_column_widths(ws, max_width: int = 20):
    """Auto-set all column widths, capped at max_width."""
    for col_idx in range(1, ws.max_column + 1):
        set_column_width(ws, col_idx, max_width)


def add_signoff_block(ws, start_row: int, title: str = "Sign-Off"):
    """Add a sign-off block (3 rows: title, blanks for name/date/signature).

    Args:
        ws: openpyxl worksheet
        start_row: Starting row number
        title: Block title (e.g. "Reviewed By", "Approved By")
    """
    from openpyxl.styles import Border, Side

    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Title row
    ws[f"A{start_row}"] = title
    ws[f"A{start_row}"].font = Font(bold=True, size=10)

    # Sign-off grid: Name | Date | Signature (3 rows)
    headers = ["Name", "Date", "Signature"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row + 1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=9)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Blank rows for entry
    for row_offset in range(2, 4):
        for col_idx in range(1, 4):
            cell = ws.cell(row=start_row + row_offset, column=col_idx, value="")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top")
            ws.row_dimensions[start_row + row_offset].height = 20


def add_reference_hyperlink(ws, row: int, col: int | str, ref_sheet: str, ref_row: int):
    """Add a hyperlink to another sheet (for cross-referencing).

    Args:
        ws: openpyxl worksheet
        row: Row number
        col: Column index (int) or letter (str)
        ref_sheet: Name of target sheet
        ref_row: Row number in target sheet
    """
    from openpyxl.utils import get_column_letter

    if isinstance(col, int):
        col = get_column_letter(col)

    cell = ws[f"{col}{row}"]
    cell.hyperlink = f"'{ref_sheet}'!A{ref_row}"
    cell.font = Font(color="0563C1", underline="single")  # Blue underlined (hyperlink style)


def apply_number_format(ws, row: int, col: int | str, format_code: str):
    """Apply a number format to a cell.

    Args:
        ws: openpyxl worksheet
        row: Row number
        col: Column index (int) or letter (str)
        format_code: Excel format code (e.g., PERCENT_FORMAT, DATE_FORMAT)
    """
    from openpyxl.utils import get_column_letter

    if isinstance(col, int):
        col = get_column_letter(col)

    ws[f"{col}{row}"].number_format = format_code


def apply_range_format(ws, start_row: int, end_row: int, col: int | str, format_code: str):
    """Apply a number format to a range of cells in a column.

    Args:
        ws: openpyxl worksheet
        start_row: Starting row number
        end_row: Ending row number (inclusive)
        col: Column index (int) or letter (str)
        format_code: Excel format code
    """
    from openpyxl.utils import get_column_letter

    if isinstance(col, int):
        col = get_column_letter(col)

    for row in range(start_row, end_row + 1):
        ws[f"{col}{row}"].number_format = format_code
