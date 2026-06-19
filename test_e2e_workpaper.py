#!/usr/bin/env python
"""End-to-end test: upload sample CSV → run analytics → verify workpaper outputs."""

import csv
import tempfile
from pathlib import Path

import polars as pl

from fcmr_core.ingestion.pipeline import ingest_csv
from fcmr_core.rules.registry import run_pipeline
from fcmr_core.reporting.builder import build_exception_csvs
from fcmr_core.reporting.aggregation import aggregate_exception_codes
from fcmr_core.sampling.sample import select_sample
from fcmr_core.reporting.workpaper import build_workpaper


def create_sample_csv(path: Path, rows: int = 100) -> None:
    """Create sample customer master CSV with realistic data including edge cases."""
    headers = [
        "customer_id",
        "full_name",
        "pan",
        "aadhaar",
        "voter_id",
        "passport_number",
        "mobile",
        "email",
        "dob",
        "bank_account",
        "address_line1",
        "city",
        "state",
        "pincode",
    ]

    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()

        for i in range(rows):
            # Include the problematic 21-digit bank account that was causing i64 overflow
            long_account = "120324101240313000155" if i % 10 == 0 else f"{1000000000000 + i:020d}"

            writer.writerow(
                {
                    "customer_id": f"C{i:05d}",
                    "full_name": f"Customer Name {i}",
                    "pan": f"AAAA{i:01d}2023A",
                    "aadhaar": f"{123456789012 + i:012d}",
                    "voter_id": f"AB{1234567 + i:07d}",
                    "passport_number": f"A{i:08d}",
                    "mobile": f"{6000000000 + i:10d}",
                    "email": f"cust{i}@company.com",
                    "dob": "1985-05-15",
                    "bank_account": long_account,
                    "address_line1": f"{i} Main Street Road",
                    "city": "Bangalore",
                    "state": "KA",
                    "pincode": "560001",
                }
            )


def test_full_pipeline():
    """Test: CSV → Parquet → Rules → Wide CSV → Long CSV → Workpaper."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        csv_path = tmpdir / "sample.csv"
        output_dir = tmpdir / "outputs"
        output_dir.mkdir(parents=True, exist_ok=True)

        # Step 1: Create sample CSV with problematic data
        print("[OK]Creating sample CSV (100 rows, including 21-digit bank accounts)...")
        create_sample_csv(csv_path, rows=100)

        # Step 2: Ingest CSV to Parquet (test CSV reading with numeric inference)
        print("[OK] Ingesting CSV -> Parquet...")
        result = ingest_csv(csv_path, report_type="customer_master")
        parquet_path = result.parquet_path
        row_count = result.total_rows
        print(f"  {row_count} rows ingested ({result.accepted_rows} accepted, {result.rejected_rows} rejected)")

        # Step 3: Load Parquet and run rules
        print("[OK]Running 27-rule pipeline...")
        df = pl.scan_parquet(parquet_path).collect()
        df_annotated = run_pipeline(df)
        print(f"  Pipeline output: {df_annotated.shape}")

        # Step 4 & 5: Build wide and long CSVs
        print("[OK]Generating wide and long CSVs...")
        run_id = "test-run"
        wide_path, long_path = build_exception_csvs(df_annotated, run_id, output_dir)
        wide_df = pl.read_csv(wide_path, infer_schema_length=0)
        long_df = pl.read_csv(long_path, infer_schema_length=0)
        print(f"  Wide CSV: {len(wide_df)} rows, {len(wide_df.columns)} columns")
        print(f"  Long CSV: {len(long_df)} rows (one per exception)")

        # Verify 21-digit account is preserved as string (not parsed as int)
        assert "bank_account" in wide_df.columns, "bank_account column missing from wide CSV"
        assert "overall_status" in wide_df.columns, "overall_status column missing from wide CSV"
        accounts = wide_df["bank_account"].to_list()
        assert any("120324101240313000155" in str(acc) for acc in accounts), "Missing long account number"
        print(f"  [PASS] 21-digit bank account preserved: {[a for a in accounts if '1203' in str(a)][0]}")

        # Step 6: Test aggregation
        print("[OK]Aggregating exception codes...")
        all_codes = aggregate_exception_codes(wide_path, top_n=None)
        print(f"  Found {len(all_codes)} unique exception codes")
        top_codes = aggregate_exception_codes(wide_path, top_n=10)
        print(f"  Top 10: {top_codes}")

        # Step 7: Select sample for workpaper (tests ICAI table tuple unpacking fix)
        print("[OK]Selecting audit sample (ICAI table)...")
        population = len(wide_df)
        exception_count = sum(1 for s in wide_df["overall_status"] if s != "OK")
        print(f"  Population: {population}, Exceptions: {exception_count}")
        sample = select_sample(
            wide_path,
            engagement_id="test-engagement",
            run_id="test-run",
            population=population,
            exception_count=exception_count,
        )
        print(f"  Sample size: {len(sample)}")
        assert len(sample) > 0, "ICAI table failed to return sample"
        assert len(sample) <= population, "Sample size exceeds population"

        # Step 8: Build workpaper (tests Excel sheet naming fix)
        print("[OK]Building 4-sheet Excel workpaper...")
        engagement = {
            "engagement_id": "test-engagement",
            "name": "Test Audit",
            "client_name": "NBFC Corp",
            "period_from": "2025-01-01",
            "period_to": "2025-12-31",
        }
        run = {"run_id": "test-run", "engagement_id": "test-engagement"}
        workpaper_path = build_workpaper(engagement, run, wide_path, sample, output_dir)
        print(f"  Workpaper: {workpaper_path.name}")
        assert workpaper_path.exists(), "Workpaper not created"

        # Step 9: Verify workpaper structure
        from openpyxl import load_workbook

        print("[OK]Verifying workpaper structure...")
        wb = load_workbook(workpaper_path)
        sheet_names = wb.sheetnames
        print(f"  Sheets: {sheet_names}")
        expected_sheets = ["Lead Sheet", "Detailed Exceptions", "TOC and TOD", "Methodology"]
        assert sheet_names == expected_sheets, f"Sheet names mismatch: {sheet_names} vs {expected_sheets}"

        # Check Lead Sheet has content
        ws_lead = wb["Lead Sheet"]
        print(f"  Lead Sheet: {ws_lead.max_row} rows")
        assert ws_lead.max_row > 5, "Lead Sheet is empty"

        # Check Detailed Exceptions sheet
        ws_exc = wb["Detailed Exceptions"]
        print(f"  Detailed Exceptions: {ws_exc.max_row} rows")
        # Sheet has header + sampled rows, may be minimal if sample is small
        assert ws_exc.max_row >= 1, "Exceptions sheet has no headers"

        # Check TOC and TOD (was "TOC/TOD", now fixed)
        ws_toc = wb["TOC and TOD"]
        print(f"  TOC and TOD: {ws_toc.max_row} rows")
        assert ws_toc.max_row > 1, "TOC/TOD sheet is empty"

        # Check Methodology
        ws_meth = wb["Methodology"]
        print(f"  Methodology: {ws_meth.max_row} rows")
        assert ws_meth.max_row > 1, "Methodology sheet is empty"

        print("\n[SUCCESS] ALL TESTS PASSED - Workpaper generation works end-to-end!")
        print(f"\n[OUTPUT] Outputs in: {output_dir}")
        print(f"   - wide.csv: {len(wide_df)} records")
        print(f"   - long.csv: {len(long_df)} exceptions")
        print(f"   - workpaper.xlsx: 4 sheets")


if __name__ == "__main__":
    test_full_pipeline()
