"""Tests for rule categorization and filtered pipeline execution."""

from __future__ import annotations

import polars as pl

from fcmr_core.rules.registry import CATEGORIES, list_categories, resolve_rule_ids, run_pipeline


def test_categories_coverage():
    """Verify all registered rules appear in exactly one category."""
    from fcmr_core.rules.registry import _ensure_rules_loaded, list_rules

    _ensure_rules_loaded()
    all_rules = {m.rule_id for m in list_rules()}

    # Collect all rule_ids from categories
    categorized_rules = set()
    for cat in CATEGORIES:
        for rule_id in cat["rule_ids"]:
            assert rule_id in all_rules, f"Unknown rule in category: {rule_id}"
            assert rule_id not in categorized_rules, f"Rule {rule_id} in multiple categories"
            categorized_rules.add(rule_id)

    # All registered rules must be in exactly one category
    assert (
        categorized_rules == all_rules
    ), f"Missing rules: {all_rules - categorized_rules}, Extra: {categorized_rules - all_rules}"


def test_list_categories():
    """Verify list_categories returns enriched category data."""
    cats = list_categories()
    assert len(cats) == 5
    assert all("id" in c and "label" in c and "rules" in c and "count" in c for c in cats)
    assert {c["id"] for c in cats} == {
        "missing_data",
        "kyc_format",
        "address_pin",
        "duplicates",
        "identity_grouping",
    }
    cat_by_id = {c["id"]: c for c in cats}
    assert cat_by_id["missing_data"]["count"] == 8  # 7 missing rules + address_completeness
    assert cat_by_id["kyc_format"]["count"] == 11
    assert (
        cat_by_id["address_pin"]["count"] == 3
    )  # pincode_exists, state_pin_match, district_pin_match
    assert cat_by_id["duplicates"]["count"] == 7
    assert cat_by_id["identity_grouping"]["count"] == 2


def test_resolve_rule_ids_empty():
    """Resolve with no selection returns None (run all)."""
    result = resolve_rule_ids([], [])
    assert result is None


def test_resolve_rule_ids_by_category():
    """Resolve by category IDs."""
    result = resolve_rule_ids(["duplicates"], [])
    assert result is not None
    assert set(result) == {
        "pan_duplicate",
        "aadhaar_duplicate",
        "mobile_duplicate",
        "bank_account_duplicate",
        "name_dob_duplicate",
        "voter_id_duplicate",
        "address_duplicate",
    }


def test_resolve_rule_ids_by_rule():
    """Resolve by individual rule IDs."""
    result = resolve_rule_ids([], ["pan_format", "pincode_exists"])
    assert result == ["pan_format", "pincode_exists"]


def test_resolve_rule_ids_union():
    """Resolve merges categories and individual rules."""
    result = resolve_rule_ids(["address_pin"], ["pan_format"])
    expected = {
        "pincode_exists",
        "state_pin_match",
        "district_pin_match",
        "pan_format",
    }
    assert set(result) == expected


def test_run_pipeline_all_rules():
    """run_pipeline with rule_ids=None runs all rules."""
    from fcmr_core.rules.registry import _ensure_rules_loaded, list_rules

    _ensure_rules_loaded()
    all_rule_ids = {m.rule_id for m in list_rules()}

    # Create minimal test dataframe
    df = pl.DataFrame(
        {
            "customer_id": ["C001"],
            "pan": ["AAAA12023A"],
            "aadhaar": ["123456789012"],
            "full_name": ["Test"],
        }
    )

    annotated = run_pipeline(df, rule_ids=None)

    # Check all rules were executed (all _exc_* columns present)
    exc_cols = {c for c in annotated.columns if c.startswith("_exc_") and c.endswith("_status")}
    rule_ids_executed = {c[5:-7] for c in exc_cols}  # Extract between "_exc_" and "_status"
    assert rule_ids_executed == all_rule_ids


def test_run_pipeline_filtered():
    """run_pipeline with rule_ids filters to selected rules only."""
    df = pl.DataFrame(
        {
            "customer_id": ["C001"],
            "pan": ["AAAA12023A"],
            "aadhaar": ["123456789012"],
            "full_name": ["Test"],
        }
    )

    # Run only duplicates category
    annotated = run_pipeline(df, rule_ids=["pan_duplicate", "aadhaar_duplicate"])

    # Check only selected rules were executed
    exc_cols = {c for c in annotated.columns if c.startswith("_exc_") and c.endswith("_status")}
    rule_ids_executed = {c[5:-7] for c in exc_cols}  # Extract between "_exc_" and "_status"
    assert rule_ids_executed == {"pan_duplicate", "aadhaar_duplicate"}


def test_run_pipeline_preserves_order():
    """run_pipeline preserves registry order even when filtered."""
    from fcmr_core.rules.registry import _ensure_rules_loaded, list_rules

    _ensure_rules_loaded()
    all_rules = list_rules()

    # Select a subset
    selected = ["pan_format", "pincode_exists", "pan_duplicate"]

    df = pl.DataFrame(
        {
            "customer_id": ["C001"],
            "pan": ["AAAA12023A"],
            "aadhaar": ["123456789012"],
            "full_name": ["Test"],
            "pincode": ["560001"],
        }
    )

    progress_calls = []

    def track_progress(completed: int, total: int, rule_id: str):
        progress_calls.append(rule_id)

    run_pipeline(df, on_progress=track_progress, rule_ids=selected)

    # Check progress calls match registry order
    expected_order = [r.rule_id for r in all_rules if r.rule_id in selected]
    assert progress_calls == expected_order


def test_run_pipeline_progress_callback():
    """on_progress callback receives correct completed/total counts."""
    df = pl.DataFrame(
        {
            "customer_id": ["C001"],
            "pan": ["AAAA12023A"],
            "aadhaar": ["123456789012"],
            "full_name": ["Test"],
        }
    )

    progress_calls = []

    def track_progress(completed: int, total: int, rule_id: str):
        progress_calls.append((completed, total, rule_id))

    # Run only 2 rules
    run_pipeline(df, on_progress=track_progress, rule_ids=["pan_format", "aadhaar_format"])

    # Check progress: completed should increment, total should be 2
    assert len(progress_calls) == 2
    assert progress_calls[0] == (1, 2, "pan_format")
    assert progress_calls[1] == (2, 2, "aadhaar_format")


def test_methodology_icfr_table_run_all():
    """Verify Methodology ICFR table includes all 5 categories on run-all."""
    import json
    import tempfile
    from pathlib import Path

    from fcmr_core.reporting.workpaper import build_workpaper
    from openpyxl import load_workbook

    # Synthetic minimal CSVs
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)

        # Create minimal wide CSV (5 records, 3 exceptions)
        wide_csv = tmpdir / "wide.csv"
        wide_csv.write_text(
            "customer_id,overall_status,exception_codes,exception_descriptions\n"
            "C001,OK,,\n"
            "C002,ERROR,PAN_DUP,PAN duplicate\n"
            "C003,OK,,\n"
            "C004,WARN,EMAIL_GENERIC,Generic email\n"
            "C005,ERROR,AADHAAR_DUP,Aadhaar duplicate\n"
        )

        # Create minimal long CSV (3 exception rows)
        long_csv = tmpdir / "long.csv"
        long_csv.write_text(
            "_row_num,customer_id,rule_id,status,exception_code,exception_description\n"
            "2,C002,pan_duplicate,ERROR,PAN_DUP,PAN duplicate\n"
            "4,C004,email_company_generic_domain,WARN,EMAIL_GENERIC,Generic email\n"
            "5,C005,aadhaar_duplicate,ERROR,AADHAAR_DUP,Aadhaar duplicate\n"
        )

        engagement = {
            "engagement_id": "TEST001",
            "name": "Test",
            "client_name": "Client",
            "period_from": "2026-01-01",
            "period_to": "2026-12-31",
        }

        # Run all (selected_rules = None)
        run = {
            "run_id": "RUN001",
            "upload_id": "UPLOAD001",
            "selected_rules": None,
        }

        upload = {
            "upload_id": "UPLOAD001",
            "filename": "test.csv",
            "row_count": 5,
            "ingested_at": "2026-06-21T10:00:00",
        }

        sample_records = []
        output_dir = tmpdir / "outputs"

        wp_path = build_workpaper(
            engagement, run, upload, wide_csv, long_csv, sample_records, output_dir
        )

        # Open workpaper and check Methodology sheet
        wb = load_workbook(wp_path)
        ws = wb["Methodology"]

        # Read all rows and look for ICFR table
        icfr_labels = []
        in_icfr = False
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0] == "ICFR Control Mapping":
                in_icfr = True
                continue
            if in_icfr and row[0] and not row[0].startswith("Category"):
                # Stop when we hit a non-header row outside the table
                if (
                    row[0] not in [
                        "Missing Data",
                        "KYC & Document Format",
                        "Address & PIN",
                        "Duplicate Detection",
                        "Identity Grouping (UCID + Beneficiary)",
                    ]
                ):
                    break
                icfr_labels.append(row[0])

        # Should have all 5 categories
        assert len(icfr_labels) == 5, f"Expected 5 ICFR rows, got {len(icfr_labels)}: {icfr_labels}"
        assert "Missing Data" in icfr_labels
        assert "KYC & Document Format" in icfr_labels
        assert "Address & PIN" in icfr_labels
        assert "Duplicate Detection" in icfr_labels
        assert "Identity Grouping (UCID + Beneficiary)" in icfr_labels


def test_methodology_icfr_table_filtered():
    """Verify Methodology ICFR table filters to selected categories only."""
    import json
    import tempfile
    from pathlib import Path

    from fcmr_core.reporting.workpaper import build_workpaper
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)

        wide_csv = tmpdir / "wide.csv"
        wide_csv.write_text(
            "customer_id,overall_status,exception_codes,exception_descriptions\n"
            "C001,OK,,\nC002,ERROR,PAN_DUP,PAN dup\n"
        )

        long_csv = tmpdir / "long.csv"
        long_csv.write_text(
            "_row_num,customer_id,rule_id,status,exception_code,exception_description\n"
            "2,C002,pan_duplicate,ERROR,PAN_DUP,PAN dup\n"
        )

        engagement = {
            "engagement_id": "TEST002",
            "name": "Test",
            "client_name": "Client",
            "period_from": "2026-01-01",
            "period_to": "2026-12-31",
        }

        # Run only duplicates category (7 rules)
        run = {
            "run_id": "RUN002",
            "upload_id": "UPLOAD002",
            "selected_rules": json.dumps([
                "pan_duplicate",
                "aadhaar_duplicate",
                "mobile_duplicate",
                "bank_account_duplicate",
                "name_dob_duplicate",
                "voter_id_duplicate",
                "address_duplicate",
            ]),
        }

        upload = {
            "upload_id": "UPLOAD002",
            "filename": "test.csv",
            "row_count": 2,
            "ingested_at": "2026-06-21T10:00:00",
        }

        sample_records = []
        output_dir = tmpdir / "outputs"

        wp_path = build_workpaper(
            engagement, run, upload, wide_csv, long_csv, sample_records, output_dir
        )

        wb = load_workbook(wp_path)
        ws = wb["Methodology"]

        icfr_labels = []
        in_icfr = False
        skip_header = False
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0] == "ICFR Control Mapping":
                in_icfr = True
                skip_header = False
                continue
            if in_icfr:
                # Skip the table header row (Category | Control Objective | Assertion | Standard)
                if row[0] == "Category" and not skip_header:
                    skip_header = True
                    continue
                if skip_header and row[0]:
                    # Collect known categories, stop at end of table
                    known_cats = {
                        "Missing Data",
                        "KYC & Document Format",
                        "Address & PIN",
                        "Duplicate Detection",
                        "Identity Grouping (UCID + Beneficiary)",
                    }
                    if row[0] not in known_cats:
                        break
                    icfr_labels.append(row[0])

        # Should have only Duplicates
        assert len(icfr_labels) == 1, f"Expected 1 ICFR row, got {len(icfr_labels)}: {icfr_labels}"
        assert "Duplicate Detection" in icfr_labels
        assert "KYC & Document Format" not in icfr_labels
        assert "Address & PIN" not in icfr_labels
