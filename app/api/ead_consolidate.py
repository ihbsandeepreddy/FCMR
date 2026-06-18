"""EAD Files consolidation — merge multiple uploaded EAD files and download."""

from __future__ import annotations

import io
import json
from datetime import datetime, timezone
from pathlib import Path

import polars as pl
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from fcmr_core.catalog import store

router = APIRouter()
_templates_dir = Path(__file__).parent.parent / "web" / "templates"
templates = Jinja2Templates(directory=str(_templates_dir))


def _build_consolidated_df(engagement_id: str | None) -> pl.DataFrame:
    """Load all ready EAD uploads for the engagement, rename to canonical columns, and stack."""
    uploads = store.list_uploads(engagement_id=engagement_id)
    ead_ready = [u for u in uploads if u["report_type"] == "ead_files" and u["status"] == "ready"]
    if not ead_ready:
        return pl.DataFrame()

    frames: list[pl.DataFrame] = []
    for upload in ead_ready:
        df = store.get_upload_df(upload["upload_id"])
        # Rename raw → canonical using stored column_mapping {raw: canonical}
        mapping: dict[str, str] = json.loads(upload.get("column_mapping") or "{}")
        rename = {raw: canonical for raw, canonical in mapping.items() if raw in df.columns}
        if rename:
            df = df.rename(rename)
        # Add source filename column so user can trace back
        df = df.with_columns(pl.lit(upload["filename"]).alias("_source_file"))
        frames.append(df)

    return pl.concat(frames, how="diagonal_relaxed")


@router.get("/dashboard/ead/consolidate", response_class=HTMLResponse)
async def ead_consolidate_page(request: Request):
    engagement_id = request.session.get("engagement_id")
    uploads = store.list_uploads(engagement_id=engagement_id)
    ead_ready = [u for u in uploads if u["report_type"] == "ead_files" and u["status"] == "ready"]
    ead_pending = [u for u in uploads if u["report_type"] == "ead_files" and u["status"] == "mapping_pending"]
    return templates.TemplateResponse(
        request=request,
        name="ead_consolidate.html",
        context={
            "ead_ready": ead_ready,
            "ead_pending": ead_pending,
            "total": len(ead_ready),
        },
    )


@router.get("/dashboard/ead/download/csv")
async def ead_download_csv(request: Request):
    engagement_id = request.session.get("engagement_id")
    df = _build_consolidated_df(engagement_id)
    if df.is_empty():
        raise HTTPException(status_code=404, detail="No ready EAD files found to consolidate.")

    csv_bytes = df.write_csv().encode("utf-8")
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"EAD_Consolidated_{timestamp}.csv"
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/dashboard/ead/download/excel")
async def ead_download_excel(request: Request):
    engagement_id = request.session.get("engagement_id")
    df = _build_consolidated_df(engagement_id)
    if df.is_empty():
        raise HTTPException(status_code=404, detail="No ready EAD files found to consolidate.")

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EAD Consolidated"

    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    data_font = Font(size=10)

    # Write headers
    cols = df.columns
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write data rows
    rows = df.to_numpy(allow_copy=True)
    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=None if (val is None or str(val) == "None") else val)
            cell.font = data_font

    # Auto-width (cap at 40)
    for ci, col in enumerate(cols, start=1):
        max_len = max(len(col), 10)
        sample = df[col].drop_nulls().head(200).cast(pl.Utf8)
        if sample.len() > 0:
            max_len = max(max_len, sample.map_elements(len, return_dtype=pl.Int32).max() or 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "EAD Consolidation Summary"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2["A3"] = "Total Rows"
    ws2["B3"] = len(df)
    ws2["A4"] = "Total Columns"
    ws2["B4"] = len(cols)
    ws2["A5"] = "Generated At"
    ws2["B5"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    if "_source_file" in df.columns:
        ws2["A7"] = "Source Files"
        ws2["A7"].font = Font(bold=True)
        source_files = df["_source_file"].unique().to_list()
        for i, fname in enumerate(source_files):
            ws2.cell(row=8 + i, column=1, value=fname)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"EAD_Consolidated_{timestamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
