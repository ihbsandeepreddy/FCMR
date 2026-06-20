"""Schema reconciliation for multi-file consolidation + consolidated-data downloads.

This replaces the old post-upload "Data Consolidation" tab and the EAD-specific
flow.  Consolidation now happens at ingest time (see ``app/api/uploads.py`` and
``fcmr_core/ingestion/consolidation.py``); when uploaded files have mismatched
header layouts, the user lands here to align each layout onto one unified column
set before the merge.  The consolidated upload's data can also be downloaded
(CSV / Excel) from its detail page via the routes at the bottom of this module.
"""

from __future__ import annotations

import io
import json
import shutil
from datetime import UTC, datetime
from pathlib import Path

import polars as pl
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from fcmr_core.catalog import store
from fcmr_core.ingestion.consolidation import (
    FileEntry,
    group_files_by_signature,
    suggest_alignment,
    unified_columns,
)

router = APIRouter()
_templates_dir = Path(__file__).parent.parent / "web" / "templates"
templates = Jinja2Templates(directory=str(_templates_dir))

_REPORT_LABELS = {
    "customer_master": "Customer Master",
    "ead_files": "EAD Files",
    "collection_report": "Collection Report",
    "disbursement_report": "Disbursement Report",
    "technical_writeoff": "Technical Write-Off",
}


def _label(report_type: str) -> str:
    return _REPORT_LABELS.get(report_type, report_type.replace("_", " ").title())


# ── Schema reconciliation ──────────────────────────────────────────────────


@router.get("/consolidate/reconcile/{batch_id}", response_class=HTMLResponse)
async def reconcile_form(request: Request, batch_id: str):
    batch = store.get_batch(batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    if batch["status"] != "reconcile_pending":
        # Already consolidated → jump to the resulting upload's mapping page.
        if batch.get("consolidated_upload_id"):
            return RedirectResponse(
                url=f"/dashboard/uploads/{batch['consolidated_upload_id']}/map-columns",
                status_code=303,
            )
        raise HTTPException(status_code=409, detail="Batch already processed")

    entries = [FileEntry.from_dict(d) for d in json.loads(batch["files_json"] or "[]")]
    groups = group_files_by_signature(entries)
    ordered = list(groups.items())  # stable order shared by GET + POST
    unified_cols = unified_columns(groups)
    alignment = suggest_alignment(groups, unified_cols)

    group_views = [
        {
            "gidx": gi,
            "signature": sig,
            "files": [f.name for f in g["files"]],
            "headers": g["headers"],
            "alignment": alignment[sig],
        }
        for gi, (sig, g) in enumerate(ordered)
    ]

    return templates.TemplateResponse(
        request=request,
        name="reconcile.html",
        context={
            "batch_id": batch_id,
            "report_type": batch["report_type"],
            "label": _label(batch["report_type"]),
            "groups": group_views,
            "unified_cols": unified_cols,
            "file_count": len(entries),
        },
    )


@router.post("/consolidate/reconcile/{batch_id}")
async def reconcile_submit(request: Request, batch_id: str):
    from app.api.uploads import _finalize_consolidated_upload

    batch = store.get_batch(batch_id)
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    if batch["status"] != "reconcile_pending":
        raise HTTPException(status_code=409, detail="Batch already processed")

    entries = [FileEntry.from_dict(d) for d in json.loads(batch["files_json"] or "[]")]
    groups = group_files_by_signature(entries)
    ordered = list(groups.items())
    n_cols = len(unified_columns(groups))

    form = await request.form()

    # Rebuild the (possibly edited) unified column names, then each group's alignment.
    unified_cols: list[str] = []
    for ci in range(n_cols):
        name = str(form.get(f"ucol_{ci}", "") or "").strip()
        unified_cols.append(name or f"column_{ci + 1}")

    alignment: dict[str, dict[str, str | None]] = {}
    for gi, (sig, _g) in enumerate(ordered):
        col_map: dict[str, str | None] = {}
        for ci, uc in enumerate(unified_cols):
            raw = str(form.get(f"align_{gi}_{ci}", "") or "").strip()
            col_map[uc] = raw if raw and raw != "__none__" else None
        alignment[sig] = col_map

    # Drop unified columns where every layout left the mapping as "— none —".
    # These would produce all-NULL columns in the output — skip them silently.
    unified_cols = [
        uc for uc in unified_cols if any(alignment[sig].get(uc) for sig in alignment)
    ]
    for sig in alignment:
        alignment[sig] = {uc: alignment[sig].get(uc) for uc in unified_cols}

    upload_id = _finalize_consolidated_upload(
        report_type=batch["report_type"],
        engagement_id=batch.get("engagement_id"),
        batch_id=batch_id,
        ordered_groups=ordered,
        alignment=alignment,
        unified_cols=unified_cols,
        source_names=[e.name for e in entries],
    )
    store.set_batch_consolidated(batch_id, upload_id)

    # Raw staged files are no longer needed once merged into the combined CSV.
    from fcmr_core.config import settings

    shutil.rmtree(settings.uploads_dir / f"_batch_{batch_id}", ignore_errors=True)

    return RedirectResponse(url=f"/dashboard/uploads/{upload_id}/map-columns", status_code=303)


# ── Consolidated-data downloads (any upload, CSV / Excel) ───────────────────


def _df_to_excel(df: pl.DataFrame, label: str) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = label[:31]

    hdr_fill = PatternFill(start_color="5C3D1E", end_color="5C3D1E", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    data_font = Font(size=10)

    cols = df.columns
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col.replace("_", " ").title())
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows_data = df.with_columns([pl.col(c).cast(pl.Utf8, strict=False) for c in cols]).rows()
    for ri, row in enumerate(rows_data, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val).font = data_font

    for ci, col in enumerate(cols, 1):
        sample = df[col].drop_nulls().head(200).cast(pl.Utf8, strict=False)
        max_len = max(len(col), 10)
        if sample.len() > 0:
            max_len = max(max_len, sample.map_elements(len, return_dtype=pl.Int32).max() or 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = f"{label} — Consolidation Summary"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2["A3"] = "Report Type"
    ws2["B3"] = label
    ws2["A4"] = "Total Rows"
    ws2["B4"] = len(df)
    ws2["A5"] = "Total Columns"
    ws2["B5"] = len(cols)
    ws2["A6"] = "Generated At"
    ws2["B6"] = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    if "_source_file" in cols:
        ws2["A8"] = "Source Files"
        ws2["A8"].font = Font(bold=True)
        for i, fname in enumerate(df["_source_file"].unique().sort().to_list()):
            ws2.cell(row=9 + i, column=1, value=fname)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _load_upload(request: Request, upload_id: str) -> tuple[dict, pl.DataFrame]:
    upload = store.get_upload(upload_id)
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    if upload["status"] != "ready":
        raise HTTPException(status_code=409, detail="Upload is not ready for download.")
    return upload, store.get_upload_df(upload_id)


@router.get("/dashboard/uploads/{upload_id}/download/csv")
async def upload_download_csv(request: Request, upload_id: str):
    upload, df = _load_upload(request, upload_id)
    ts = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    filename = f"{_label(upload['report_type']).replace(' ', '_')}_{ts}.csv"
    return StreamingResponse(
        io.BytesIO(df.write_csv().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/dashboard/uploads/{upload_id}/download/excel")
async def upload_download_excel(request: Request, upload_id: str):
    upload, df = _load_upload(request, upload_id)
    buf = _df_to_excel(df, _label(upload["report_type"]))
    ts = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    filename = f"{_label(upload['report_type']).replace(' ', '_')}_{ts}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
