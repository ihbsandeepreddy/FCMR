"""Download endpoints for exception CSVs and missing-data Excel."""

import io
from pathlib import Path

import polars as pl
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from fcmr_core.catalog import store
from fcmr_core.reporting.aggregation import _MISSING_CODES, _MISSING_LABELS

router = APIRouter()


@router.get("/runs/{run_id}/download/wide")
async def download_wide(run_id: str):
    run = store.get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if not run.get("wide_csv"):
        raise HTTPException(status_code=404, detail="Wide CSV not available")
    p = Path(run["wide_csv"])
    if not p.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    return FileResponse(
        path=str(p),
        media_type="text/csv",
        filename=f"{run_id}_exceptions_wide.csv",
    )


@router.get("/runs/{run_id}/download/long")
async def download_long(run_id: str):
    run = store.get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if not run.get("long_csv"):
        raise HTTPException(status_code=404, detail="Long CSV not available")
    p = Path(run["long_csv"])
    if not p.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    return FileResponse(
        path=str(p),
        media_type="text/csv",
        filename=f"{run_id}_exceptions_long.csv",
    )


@router.get("/runs/{run_id}/download/missing-data")
async def download_missing_data(run_id: str):
    """Build and download a 2-sheet Missing Data Excel workbook."""
    run = store.get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Run is not completed")
    if not run.get("long_csv"):
        raise HTTPException(status_code=404, detail="Long CSV not available")

    long_path = Path(run["long_csv"])
    if not long_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")

    wide_path = Path(run["wide_csv"]) if run.get("wide_csv") else None
    total_rows = 0
    if wide_path and wide_path.exists():
        try:
            wdf = pl.read_csv(wide_path, columns=["overall_status"], infer_schema_length=0)
            total_rows = len(wdf)
        except Exception:
            pass

    # Read long CSV and filter to missing codes
    try:
        long_df = pl.read_csv(long_path, infer_schema_length=0)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to read long CSV: {exc}") from exc

    missing_codes_col = "exception_code" if "exception_code" in long_df.columns else None
    if missing_codes_col is None:
        raise HTTPException(status_code=500, detail="exception_code column not found in long CSV")

    missing_df = long_df.filter(pl.col("exception_code").is_in(list(_MISSING_CODES)))

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "Summary"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    summary_headers = ["Field", "Exception Code", "Missing Count", "% of Total"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Count per code
    code_counts: dict[str, int] = {}
    for code in missing_df["exception_code"]:
        if code:
            code_counts[code] = code_counts.get(code, 0) + 1

    row = 2
    for code, cnt in sorted(code_counts.items(), key=lambda x: x[1], reverse=True):
        pct = round(cnt / total_rows * 100, 1) if total_rows > 0 else 0.0
        ws1.cell(row=row, column=1, value=_MISSING_LABELS.get(code, code))
        ws1.cell(row=row, column=2, value=code)
        ws1.cell(row=row, column=3, value=cnt)
        ws1.cell(row=row, column=4, value=f"{pct}%")
        row += 1

    for col_idx, width in enumerate([25, 25, 15, 12], 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Sheet 2: Detail ──
    ws2 = wb.create_sheet("Detail")
    detail_cols = [
        c
        for c in ["_row_num", "customer_id", "exception_code", "exception_description"]
        if c in missing_df.columns
    ]
    if not detail_cols:
        detail_cols = missing_df.columns[:4]

    for col_idx, h in enumerate(detail_cols, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font

    detail_sub = missing_df.select(detail_cols)
    for row_idx, data_row in enumerate(detail_sub.rows(), 2):
        for col_idx, val in enumerate(data_row, 1):
            ws2.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(detail_cols) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{run_id}_missing_data.xlsx"',
        },
    )
